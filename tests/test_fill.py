import os
import warnings

import openpyxl

from src.myntra.template_reader import read_template
from src.core.models import MappedRow, ImageResult
from src.myntra.fill import fill_template

TEMPLATE = "templates/myntra/Myntra-Sku-Template-2026-06-16.xlsx"


def test_fill_writes_rows(tmp_path):
    warnings.filterwarnings("ignore")
    t = read_template(TEMPLATE)
    r1 = MappedRow(sku="S1", cells={"vendorSkuCode": "S1", "vendorArticleName": "Blue Saree",
                                    "MRP": "3499", "ISP": "3199", "articleType": "Sarees"})
    img = ImageResult(sku="S1", jpgs=["S1_1.jpg"], passed=["/x/S1_1.jpg"],
                      passed_urls=["https://cdn.shopify.com/x/Blue-1.webp?v=1"], failed=[])
    out = tmp_path / "filled.xlsx"
    fill_template(TEMPLATE, t, [(r1, img)], str(out))
    assert os.path.exists(out)
    wb = openpyxl.load_workbook(out)
    ws = wb["Sarees"]
    row = t.first_data_row
    sku_col = t.col_index_by_header["vendorSkuCode"]
    name_col = t.col_index_by_header["vendorArticleName"]
    front_col = t.col_index_by_header["Front Image"]
    assert ws.cell(row=row, column=sku_col).value == "S1"
    assert ws.cell(row=row, column=name_col).value == "Blue Saree"
    assert ws.cell(row=row, column=front_col).value == "https://cdn.shopify.com/x/Blue-1.webp?v=1"


def test_template_example_rows_are_cleared(tmp_path):
    """The template ships with leftover example image URLs (e.g. row 11). After a
    fill of N products, no row beyond the data must carry stray content."""
    warnings.filterwarnings("ignore")
    t = read_template(TEMPLATE)
    r1 = MappedRow(sku="S1", cells={"vendorSkuCode": "S1", "brand": "Ijor Ethnic Partners"})
    out = tmp_path / "filled.xlsx"
    fill_template(TEMPLATE, t, [(r1, ImageResult(sku="S1"))], str(out))
    ws = openpyxl.load_workbook(out)["Sarees"]
    # Only the single product row (row 4) should hold data; row 11's template
    # example URLs must be gone.
    rows_with_data = [r for r in range(t.first_data_row, ws.max_row + 1)
                      if any(ws.cell(r, c).value not in (None, "") for c in range(1, 81))]
    assert rows_with_data == [t.first_data_row]
    assert ws.cell(11, 74).value is None
