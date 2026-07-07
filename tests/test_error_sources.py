import csv
import openpyxl
from src.myntra.error_sources import detect_format, read_error_file, ErrorItem
from src.myntra.error_reader import load_rules


def _sku_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sarees"
    headers = ["STATUS", "SYSTEM ERROR MESSAGE", "styleGroupId", "vendorSkuCode"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    ws.cell(row=4, column=1, value="SKU_VALIDATION_FAILED")
    ws.cell(row=4, column=2, value="ISP cannot be empty; 6 digit Pincode is missing")
    ws.cell(row=4, column=3, value=11)
    ws.cell(row=4, column=4, value="78SAZ125BSI")
    wb.save(path)


def _sheet_csv(path):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["ROW NO", "BRAND", "STATUS", "SYSTEM ERROR MESSAGE"])
        w.writerow(["0", "", "SHEET_VALIDATION_FAILED",
                    "Style SKU Count Validation failed! : Minimum unique StyleGroupIds required is 1. Given sheet has only 7."])


def _listings_csv(path):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["article type", "brand", "style status", "style id",
                    "seller sku code", "onhold reason"])
        w.writerow(["Sarees", "Ijor", "P", "43214808", "127SDE826NSB", ""])           # live -> skipped
        w.writerow(["Sarees", "Ijor", "PMR", "43214809", "128SDE826NSB",
                    "Image is a flat shot; reshoot on model"])                          # rejected


def test_detect_sku_xlsx(tmp_path):
    p = tmp_path / "r.xlsx"
    _sku_xlsx(p)
    src, reason = detect_format(str(p))
    assert src == "sku_xlsx"


def test_detect_sheet_csv(tmp_path):
    p = tmp_path / "e.csv"
    _sheet_csv(p)
    assert detect_format(str(p))[0] == "sheet_csv"


def test_detect_listings_report(tmp_path):
    p = tmp_path / "l.csv"
    _listings_csv(p)
    assert detect_format(str(p))[0] == "listings_report"


def test_detect_unknown_extension(tmp_path):
    p = tmp_path / "x.txt"
    p.write_text("hi", encoding="utf-8")
    src, reason = detect_format(str(p))
    assert src is None
    assert reason


def test_read_sku_xlsx_splits_clauses(tmp_path):
    p = tmp_path / "r.xlsx"
    _sku_xlsx(p)
    items = read_error_file(str(p), load_rules())
    assert all(isinstance(i, ErrorItem) for i in items)
    assert {i.raw_reason for i in items} == {
        "ISP cannot be empty", "6 digit Pincode is missing"}
    assert all(i.source_type == "sku_xlsx" and i.scope == "sku" for i in items)
    assert items[0].cells["vendorSkuCode"] == "78SAZ125BSI"


def test_read_sheet_csv_one_item_no_split(tmp_path):
    p = tmp_path / "e.csv"
    _sheet_csv(p)
    items = read_error_file(str(p))
    assert len(items) == 1
    assert items[0].scope == "sheet"
    assert items[0].sku is None
    assert "Style SKU Count" in items[0].raw_reason


def test_read_listings_skips_live_rows(tmp_path):
    p = tmp_path / "l.csv"
    _listings_csv(p)
    items = read_error_file(str(p))
    assert len(items) == 1
    assert items[0].sku == "128SDE826NSB"
    assert items[0].style_id == "43214809"
    assert items[0].source_type == "listings_report"
