import io
import warnings
import openpyxl
from PIL import Image
from src.myntra.pipeline import main


def test_default_template_is_v13_and_has_vocab():
    from src.myntra.pipeline import DEFAULT_TEMPLATE_NAME, _resolve
    from src.myntra.template_reader import read_template
    assert DEFAULT_TEMPLATE_NAME == "Myntra-Sku-Template-2026-07-24.xlsx"
    path = _resolve(DEFAULT_TEMPLATE_NAME, "templates/myntra")
    t = read_template(path)
    assert "Banarasi" in t.vocab_by_header["Type"]


def _fake_fetch():
    buf = io.BytesIO()
    Image.new("RGB", (1000, 1000), (10, 20, 30)).save(buf, "PNG")
    data = buf.getvalue()
    return lambda url: data


def test_style_group_id_start_override(tmp_path):
    warnings.filterwarnings("ignore")
    out = tmp_path / "out"
    main(
        template_path="templates/myntra/Myntra-Sku-Template-2026-06-16.xlsx",
        csv_path="tests/fixtures/products_export.csv",
        out_dir=str(out),
        config_dir="config/myntra",
        fetch=_fake_fetch(),
        upload=False,
        style_group_id_start=100,
    )
    ws = openpyxl.load_workbook(out / "myntra_filled.xlsx")["Sarees"]
    hdr = {ws.cell(3, c).value: c for c in range(1, ws.max_column + 1)}
    # fixture has 2 products -> styleGroupIds 100, 101
    assert ws.cell(4, hdr["styleGroupId"]).value == 100
    assert ws.cell(5, hdr["styleGroupId"]).value == 101


def _csv_with_hsn(tmp_path, rows):
    """The shared fixture's columns plus the HSN metafield column.

    The shared fixture itself must NOT gain an HSN column — test_shopify_reader
    and test_end_to_end both read it — so batches that need one write their own."""
    from src.core.shopify_reader import HSN_COL
    path = tmp_path / "export.csv"
    header = ("Handle,Title,Vendor,Tags,Body (HTML),Variant SKU,Variant Price,"
              "Variant Compare At Price,Image Src,Image Position,Status,"
              "Color (product.metafields.shopify.color-pattern),"
              "Fabric (product.metafields.shopify.fabric),"
              "Size (product.metafields.shopify.size)," + HSN_COL + "\n")
    path.write_text(header + "".join(rows), encoding="utf-8")
    return str(path)


def test_export_hsn_written_to_sheet(tmp_path):
    warnings.filterwarnings("ignore")
    out = tmp_path / "out"
    csv = _csv_with_hsn(tmp_path, [
        "h1,Cotton Saree,V,,,S1,1200,1500,https://example.com/a.webp,1,active,red,cotton,Free,52081120\n",
        "h2,Silk Saree,V,,,S2,2500,,https://example.com/b.webp,1,active,blue,silk,Free,50072010 \n",
    ])
    main(template_path="templates/myntra/Myntra-Sku-Template-2026-06-16.xlsx",
         csv_path=csv, out_dir=str(out), config_dir="config/myntra",
         fetch=_fake_fetch(), upload=False)
    ws = openpyxl.load_workbook(out / "myntra_filled.xlsx")["Sarees"]
    hdr = {ws.cell(3, c).value: c for c in range(1, ws.max_column + 1)}
    # HSN is in NUMERIC_HEADERS -> written as an integer cell
    assert ws.cell(4, hdr["HSN"]).value == 52081120
    assert ws.cell(5, hdr["HSN"]).value == 50072010     # trailing space normalised


def test_missing_or_malformed_hsn_leaves_the_cell_blank(tmp_path):
    warnings.filterwarnings("ignore")
    out = tmp_path / "out"
    csv = _csv_with_hsn(tmp_path, [
        "h1,Cotton Saree,V,,,S1,1200,1500,https://example.com/a.webp,1,active,red,cotton,Free,\n",
        "h2,Silk Saree,V,,,S2,2500,,https://example.com/b.webp,1,active,blue,silk,Free,5407\n",
    ])
    main(template_path="templates/myntra/Myntra-Sku-Template-2026-06-16.xlsx",
         csv_path=csv, out_dir=str(out), config_dir="config/myntra",
         fetch=_fake_fetch(), upload=False)
    ws = openpyxl.load_workbook(out / "myntra_filled.xlsx")["Sarees"]
    hdr = {ws.cell(3, c).value: c for c in range(1, ws.max_column + 1)}
    assert ws.cell(4, hdr["HSN"]).value in (None, "")   # metafield not filled
    assert ws.cell(5, hdr["HSN"]).value in (None, "")   # 4 digits, not usable


def test_content_hash_is_unaffected_by_hsn(tmp_path):
    """The duplicate-generation guard must not notice HSN.

    sku_registry._EXCLUDE already drops it from the fingerprint. If that ever
    changed, every SKU already in the registry would hash differently and the
    next upload would report the whole catalogue as "edited"."""
    from src.myntra.pipeline import scan_content_hashes
    (tmp_path / "a").mkdir()
    (tmp_path / "b").mkdir()
    row = ("h1,Cotton Saree,V,,,S1,1200,1500,https://example.com/a.webp,1,"
           "active,red,cotton,Free,")
    with_hsn = _csv_with_hsn(tmp_path / "b", [row + "52081120\n"])
    without = _csv_with_hsn(tmp_path / "a", [row + "\n"])
    assert scan_content_hashes(without) == scan_content_hashes(with_hsn)


def test_hsn_override_wins_in_mapper():
    from src.myntra.mapper import map_product
    from src.core.models import Product, TemplateInfo
    headers = ["SKUCode", "HSN"]
    tmpl = TemplateInfo(headers=headers, header_row=3, first_data_row=4,
                        col_index_by_header={h: i + 1 for i, h in enumerate(headers)},
                        vocab_by_header={})
    p = Product(handle="h", sku="S1", title="T", vendor="", tags="", body_html="",
                price=1.0, compare_at_price=None, color=None, fabric="Pure Silk",
                size=None, status="active", images=[])
    row = map_product(p, tmpl, {}, {"articleType": "Sarees"}, {},
                      hsn="50072010",
                      hsn_override="99999999")
    assert row.cells["HSN"] == "99999999"


def test_scan_content_hashes_pairs_sku_and_hash():
    from src.myntra.pipeline import scan_content_hashes
    pairs = scan_content_hashes("tests/fixtures/products_export.csv",
                                template_path="templates/myntra/Myntra-Sku-Template-2026-06-16.xlsx")
    assert len(pairs) == 2
    skus = [s for s, _ in pairs]
    assert len(set(skus)) == 2                 # distinct SKUs
    assert all(len(h) == 40 for _, h in pairs)  # sha1 hex


def test_pipeline_pins_id_hsn_and_returns_records(tmp_path):
    warnings.filterwarnings("ignore")
    from src.myntra.pipeline import main, scan_content_hashes
    pairs = dict(scan_content_hashes("tests/fixtures/products_export.csv",
                 template_path="templates/myntra/Myntra-Sku-Template-2026-06-16.xlsx"))
    sku0 = list(pairs)[0]
    res = main(
        template_path="templates/myntra/Myntra-Sku-Template-2026-06-16.xlsx",
        csv_path="tests/fixtures/products_export.csv",
        out_dir=str(tmp_path / "out"), config_dir="config/myntra",
        fetch=_fake_fetch(), upload=False,
        only_skus={sku0},
        style_group_id_by_sku={sku0: 77},
        hsn_by_sku={sku0: "63079090"},
    )
    assert res["products"] == 1                         # filtered to one SKU
    rec = res["records"][0]
    assert rec["sku"] == sku0
    assert rec["style_group_id"] == 77
    assert rec["hsn"] == "63079090"
    assert rec["content_hash"] == pairs[sku0]           # excludes id+HSN → matches scan
    ws = openpyxl.load_workbook(tmp_path / "out" / "myntra_filled.xlsx")["Sarees"]
    hdr = {ws.cell(3, c).value: c for c in range(1, ws.max_column + 1)}
    assert ws.cell(4, hdr["styleGroupId"]).value == 77
    assert ws.cell(4, hdr["HSN"]).value == 63079090


def test_pipeline_pins_names_into_the_built_sheet(tmp_path):
    """The end of the pin chain: a name pinned in the registry must survive a
    rebuild and land in the workbook, beating the title-derived default."""
    warnings.filterwarnings("ignore")
    from src.myntra.pipeline import main, scan_content_hashes
    template = "templates/myntra/Myntra-Sku-Template-2026-07-24.xlsx"
    pairs = dict(scan_content_hashes("tests/fixtures/products_export.csv",
                                     template_path=template))
    sku0 = list(pairs)[0]
    res = main(
        template_path=template,
        csv_path="tests/fixtures/products_export.csv",
        out_dir=str(tmp_path / "out"), config_dir="config/myntra",
        fetch=_fake_fetch(), upload=False,
        only_skus={sku0},
        names_by_sku={sku0: {"productDisplayName": "Ijor Handloom Saree",
                             "List View Name": "Ijor Saree"}},
    )
    assert res["products"] == 1
    ws = openpyxl.load_workbook(tmp_path / "out" / "myntra_filled.xlsx")["Sarees"]
    hdr = {ws.cell(3, c).value: c for c in range(1, ws.max_column + 1)}
    assert ws.cell(4, hdr["productDisplayName"]).value == "Ijor Handloom Saree"
    assert ws.cell(4, hdr["List View Name"]).value == "Ijor Saree"


def test_pipeline_without_a_name_pin_still_writes_the_shopify_title(tmp_path):
    warnings.filterwarnings("ignore")
    from src.myntra.pipeline import main
    template = "templates/myntra/Myntra-Sku-Template-2026-07-24.xlsx"
    res = main(
        template_path=template,
        csv_path="tests/fixtures/products_export.csv",
        out_dir=str(tmp_path / "out"), config_dir="config/myntra",
        fetch=_fake_fetch(), upload=False,
    )
    assert res["products"] >= 1
    ws = openpyxl.load_workbook(tmp_path / "out" / "myntra_filled.xlsx")["Sarees"]
    hdr = {ws.cell(3, c).value: c for c in range(1, ws.max_column + 1)}
    assert ws.cell(4, hdr["productDisplayName"]).value        # title-derived
    assert ws.cell(4, hdr["List View Name"]).value is None    # deliberately blank
