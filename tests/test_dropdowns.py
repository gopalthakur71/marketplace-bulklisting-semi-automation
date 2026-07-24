import re
import warnings
import zipfile

import openpyxl

from src.myntra.template_reader import read_template
from src.core.models import MappedRow, ImageResult
from src.myntra.fill import fill_template

TEMPLATE = "templates/myntra/Myntra-Sku-Template-2026-06-16.xlsx"
V13 = "templates/myntra/Myntra-Sku-Template-2026-07-24.xlsx"


def _count_x14_validations(xlsx_path):
    with zipfile.ZipFile(xlsx_path) as z:
        xml = z.read("xl/worksheets/sheet2.xml").decode("utf-8")
    return len(re.findall(r"<x14:dataValidation\b", xml))


def test_output_preserves_dropdowns(tmp_path):
    warnings.filterwarnings("ignore")
    original = _count_x14_validations(TEMPLATE)
    assert original == 37
    t = read_template(TEMPLATE)
    row = MappedRow(sku="S1", cells={"vendorSkuCode": "S1", "brand": "Ijor Ethnic Partners"})
    img = ImageResult(sku="S1")
    out = tmp_path / "filled.xlsx"
    fill_template(TEMPLATE, t, [(row, img)], str(out), preserve_dropdowns=True)
    assert _count_x14_validations(str(out)) == 37


def test_upload_file_has_no_dropdowns_by_default(tmp_path):
    """Default output must be clean (no x14 validations) so Myntra's parser reads it."""
    warnings.filterwarnings("ignore")
    t = read_template(TEMPLATE)
    row = MappedRow(sku="S1", cells={"vendorSkuCode": "S1"})
    out = tmp_path / "filled.xlsx"
    fill_template(TEMPLATE, t, [(row, ImageResult(sku="S1"))], str(out))
    assert _count_x14_validations(str(out)) == 0


def _count_plain_validations(path):
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(path)
    n = len(wb["Sarees"].data_validations.dataValidation)
    wb.close()
    return n


def test_v13_output_preserves_plain_dropdowns_and_blanks_attributes(tmp_path):
    warnings.filterwarnings("ignore")
    assert _count_plain_validations(V13) == 11100
    t = read_template(V13)
    row = MappedRow(sku="S1", cells={"vendorSkuCode": "S1", "brand": "Ijor Ethnic Partners"})
    out = tmp_path / "filled.xlsx"
    fill_template(V13, t, [(row, ImageResult(sku="S1"))], str(out))
    # dropdowns survive fill_template (incl. its shared-string -> inline pass)
    assert _count_plain_validations(str(out)) == 11100
    # the 9 attribute columns are empty in the data row
    wb = openpyxl.load_workbook(str(out))
    ws = wb["Sarees"]
    for header in ["Prominent Colour", "Saree Fabric", "Blouse Fabric", "Type",
                   "Ornamentation", "Border", "Pattern", "Print or Pattern Type",
                   "Wash Care"]:
        col = t.col_index_by_header[header]
        assert ws.cell(row=t.first_data_row, column=col).value in (None, "")
    wb.close()
