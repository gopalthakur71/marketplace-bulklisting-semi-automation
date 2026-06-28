# tests/test_error_reader.py
import openpyxl
from src.myntra.error_reader import load_rules, classify, read_errors


def _make_resub(path, rows):
    """Build a minimal Myntra resubmission xlsx. `rows` = list of
    (status, message, stylegroupid, vendorSkuCode)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sarees"
    headers = ["STATUS", "SYSTEM ERROR MESSAGE", "styleGroupId", "vendorSkuCode"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    r = 4
    for status, msg, sgid, sku in rows:
        ws.cell(row=r, column=1, value=status)
        ws.cell(row=r, column=2, value=msg)
        ws.cell(row=r, column=3, value=sgid)
        ws.cell(row=r, column=4, value=sku)
        r += 1
    wb.save(path)


def test_classify_known_and_unknown():
    rules = load_rules()
    dup = classify("Seller Sku Code X is already registered for seller 87065", rules)
    assert dup["category"] == "duplicate"
    assert dup["action"] == "drop_sku"
    unk = classify("some brand new error wording", rules)
    assert unk["category"] == "unknown"
    assert unk["action"] == "explain_only"


def test_read_errors_parses_rows_and_issues(tmp_path):
    p = tmp_path / "resub.xlsx"
    _make_resub(p, [
        ("SKU_VALIDATION_FAILED",
         "ISP cannot be empty for DIY source.; 6 digit Pincode is missing in manufacturer name and address",
         11, "78SAZ125BSI"),
        ("SKU_VALIDATION_FAILED",
         "Seller Sku Code 165SDE226RSG is already registered for seller 87065",
         12, "165SDE226RSG"),
    ])
    rules = load_rules()
    errs = read_errors(str(p), rules)
    assert len(errs) == 2
    first = errs[0]
    assert first.sku == "78SAZ125BSI"
    assert first.cells["styleGroupId"] == "11"          # values returned as strings
    assert "STATUS" not in first.cells                  # error columns stripped
    cats = {i["category"] for i in first.issues}
    assert cats == {"numeric", "pincode"}               # two ;-separated messages classified
    assert errs[1].issues[0]["action"] == "drop_sku"
