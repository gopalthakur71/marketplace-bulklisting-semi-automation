import warnings
import zipfile

import openpyxl
import pytest

from src.core.models import MappedRow, ImageResult
from src.myntra.attribute_entry import (user_filled_attributes, attribute_vocab,
                                        validate_submitted, AttributeValueError,
                                        write_attributes, SkuMismatchError)
from src.myntra.fill import fill_template
from src.myntra.pipeline import DEFAULT_TEMPLATE_NAME, _resolve
from src.myntra.template_reader import read_template

V13 = _resolve(DEFAULT_TEMPLATE_NAME, "templates/myntra")


def test_user_filled_attributes_reads_the_twelve_from_rules():
    cols = user_filled_attributes()
    assert len(cols) == 12
    assert cols[0] == "Prominent Colour"
    assert "Usage" in cols


def test_attribute_vocab_only_from_template_and_na_where_the_sheet_has_it():
    t = read_template(V13)
    vocab = attribute_vocab(t, user_filled_attributes())
    assert set(vocab) == set(user_filled_attributes())
    # exact sizes read off V13 (2026-07-27)
    assert len(vocab["Prominent Colour"]) == 53
    assert len(vocab["Usage"]) == 9
    assert len(vocab["Border"]) == 10
    # NA is offered only where Myntra actually lists it — never injected
    for col in ["Prominent Colour", "Second Prominent Colour", "Third Prominent Colour",
                "Blouse Fabric", "Type", "Ornamentation", "Usage"]:
        assert any(v.strip().upper() == "NA" for v in vocab[col]), col
    for col in ["Saree Fabric", "Border", "Pattern", "Print or Pattern Type", "Wash Care"]:
        assert not any(v.strip().upper() == "NA" for v in vocab[col]), col


def test_validate_submitted_blank_becomes_none():
    vocab = {"Border": ["Zari", "Solid"], "Pattern": ["Solid"]}
    out = validate_submitted({"Border": "", "Pattern": "   "}, vocab)
    assert out == {"Border": None, "Pattern": None}


def test_validate_submitted_passes_exact_vocab_values():
    vocab = {"Border": ["Zari", "Solid"]}
    assert validate_submitted({"Border": "Zari"}, vocab) == {"Border": "Zari"}


def test_validate_submitted_rejects_off_vocab_value():
    vocab = {"Border": ["Zari", "Solid"]}
    with pytest.raises(AttributeValueError) as exc:
        validate_submitted({"Border": "Salmon Pink"}, vocab)
    assert "Border" in str(exc.value)
    assert "Salmon Pink" in str(exc.value)


def test_derive_brand_colour_mirrors_the_prominent_colour_lowercased():
    from src.myntra.attribute_entry import derive_brand_colour
    assert derive_brand_colour({"Prominent Colour": "Blue"}) == "blue"


def test_derive_brand_colour_is_none_when_no_colour_was_chosen():
    from src.myntra.attribute_entry import derive_brand_colour
    assert derive_brand_colour({"Prominent Colour": None}) is None
    assert derive_brand_colour({"Prominent Colour": "  "}) is None
    assert derive_brand_colour({}) is None


def test_derive_brand_colour_treats_na_as_no_colour():
    """NA is a real member of the Prominent Colour vocabulary, so a plain
    truthiness check would write the string 'na' into the sheet."""
    from src.myntra.attribute_entry import derive_brand_colour
    assert derive_brand_colour({"Prominent Colour": "NA"}) is None


def test_validate_submitted_rejects_unknown_column():
    with pytest.raises(AttributeValueError):
        validate_submitted({"Nonexistent Column": "x"}, {"Border": ["Zari"]})


def _built(tmp_path, skus=("S1", "S2")):
    """A freshly built workbook: identity columns filled, the 12 attrs blank."""
    warnings.filterwarnings("ignore")
    t = read_template(V13)
    rows = [(MappedRow(sku=s, cells={"vendorSkuCode": s, "brand": "Ijor"}),
             ImageResult(sku=s)) for s in skus]
    out = tmp_path / "myntra_filled.xlsx"
    fill_template(V13, t, rows, str(out))
    return t, str(out)


def _validation_count(path):
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(path)
    n = len(wb["Sarees"].data_validations.dataValidation)
    wb.close()
    return n


def _sarees_xml(path):
    from src.myntra.fill import sheet_xml_name
    with zipfile.ZipFile(path) as z:
        return z.read(sheet_xml_name(path, "Sarees")).decode("utf-8")


def _cell(path, template, row_ordinal, header):
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(path, data_only=True)
    v = wb["Sarees"].cell(row=template.first_data_row + row_ordinal,
                          column=template.col_index_by_header[header]).value
    wb.close()
    return v


def test_write_attributes_writes_the_right_row_and_leaves_others_blank(tmp_path):
    t, path = _built(tmp_path)
    n = write_attributes(path, t, [
        {"ordinal": 1, "sku": "S2",
         "values": {"Border": "Zari", "Type": "Banarasi", "Pattern": None}}])
    assert n == 1
    assert _cell(path, t, 1, "Border") == "Zari"
    assert _cell(path, t, 1, "Type") == "Banarasi"
    assert _cell(path, t, 1, "Pattern") is None       # explicit blank stays blank
    assert _cell(path, t, 0, "Border") is None        # the other SKU is untouched
    assert _cell(path, t, 1, "vendorSkuCode") == "S2"  # identity columns preserved


def test_write_attributes_is_idempotent_and_can_clear(tmp_path):
    t, path = _built(tmp_path)
    write_attributes(path, t, [{"ordinal": 0, "sku": "S1", "values": {"Border": "Zari"}}])
    write_attributes(path, t, [{"ordinal": 0, "sku": "S1", "values": {"Border": "Solid"}}])
    assert _cell(path, t, 0, "Border") == "Solid"
    write_attributes(path, t, [{"ordinal": 0, "sku": "S1", "values": {"Border": None}}])
    assert _cell(path, t, 0, "Border") is None


def test_write_attributes_preserves_dropdowns(tmp_path):
    """KEY INVARIANT: the downloaded file must still have live Excel dropdowns."""
    t, path = _built(tmp_path)
    before = _validation_count(path)
    assert before > 0
    write_attributes(path, t, [{"ordinal": 0, "sku": "S1", "values": {"Border": "Zari"}}])
    assert _validation_count(path) == before


def test_write_attributes_keeps_strings_inline(tmp_path):
    """Myntra's parser cannot resolve shared strings; a bare openpyxl save undoes
    fill.py's inline conversion, so the save path must re-apply it."""
    t, path = _built(tmp_path)
    write_attributes(path, t, [{"ordinal": 0, "sku": "S1", "values": {"Border": "Zari"}}])
    xml = _sarees_xml(path)
    assert 't="s"' not in xml
    assert "Zari" in xml


def test_write_attributes_rejects_sku_mismatch_without_writing(tmp_path):
    t, path = _built(tmp_path)
    with pytest.raises(SkuMismatchError):
        write_attributes(path, t, [
            {"ordinal": 0, "sku": "S1", "values": {"Border": "Zari"}},
            {"ordinal": 1, "sku": "WRONG", "values": {"Border": "Solid"}}])
    assert _cell(path, t, 0, "Border") is None   # nothing written at all


def test_user_filled_freetext_reads_the_yaml_list():
    from src.myntra.attribute_entry import user_filled_freetext
    assert user_filled_freetext() == ["tags", "List View Name", "productDisplayName"]


def test_the_two_name_columns_are_free_text_not_dropdowns():
    """Neither name has a template vocabulary, so validating them by exact
    membership would reject every value the seller could possibly type."""
    from src.myntra.attribute_entry import (user_filled_attributes,
                                            user_filled_freetext)
    for column in ("List View Name", "productDisplayName"):
        assert column in user_filled_freetext()
        assert column not in user_filled_attributes()


def test_validate_freetext_accepts_a_typed_product_display_name():
    from src.myntra.attribute_entry import validate_freetext
    columns = ["tags", "List View Name", "productDisplayName"]
    out = validate_freetext(
        {"productDisplayName": "Ijor Handloom Cotton Saree"}, columns)
    assert out == {"productDisplayName": "Ijor Handloom Cotton Saree"}


def test_validate_freetext_accepts_any_value():
    from src.myntra.attribute_entry import validate_freetext
    out = validate_freetext({"tags": "saree, cotton, handloom"}, ["tags"])
    assert out == {"tags": "saree, cotton, handloom"}


def test_validate_freetext_strips_whitespace():
    from src.myntra.attribute_entry import validate_freetext
    assert validate_freetext({"tags": "  festive  "}, ["tags"]) == {"tags": "festive"}


def test_validate_freetext_turns_blank_into_none():
    from src.myntra.attribute_entry import validate_freetext
    assert validate_freetext({"tags": "   "}, ["tags"]) == {"tags": None}
    assert validate_freetext({"tags": ""}, ["tags"]) == {"tags": None}
    assert validate_freetext({"tags": None}, ["tags"]) == {"tags": None}


def test_validate_freetext_rejects_an_unknown_column():
    from src.myntra.attribute_entry import AttributeValueError, validate_freetext
    with pytest.raises(AttributeValueError):
        validate_freetext({"styleGroupId": "9999"}, ["tags"])


def test_validate_freetext_does_not_check_any_vocabulary():
    """The whole point: a value that would be rejected as a dropdown is fine here."""
    from src.myntra.attribute_entry import validate_freetext
    assert validate_freetext({"tags": "Salmon Pink"}, ["tags"]) == {"tags": "Salmon Pink"}


def test_validate_hsn_accepts_eight_digits_and_strips():
    from src.myntra.attribute_entry import validate_hsn
    assert validate_hsn("54075240") == "54075240"
    assert validate_hsn(" 52085990 ") == "52085990"


def test_validate_hsn_blank_clears_the_cell():
    from src.myntra.attribute_entry import validate_hsn
    assert validate_hsn("") is None
    assert validate_hsn("   ") is None
    assert validate_hsn(None) is None


def test_validate_hsn_rejects_a_non_empty_bad_value():
    from src.myntra.attribute_entry import AttributeValueError, validate_hsn
    for bad in ("5407", "6211.42.90", "abc", "540752401"):
        with pytest.raises(AttributeValueError):
            validate_hsn(bad)


def test_write_attributes_stores_hsn_as_a_number(tmp_path):
    """HSN is in fill.NUMERIC_HEADERS. A known-good upload stores those columns as
    real numeric cells — Myntra rejects text there as "non numeric" — and the build
    path coerces them. This save path must coerce too, or an HSN typed on the
    attribute screen silently uploads as text while the same code from the export
    uploads as a number."""
    t, path = _built(tmp_path)
    write_attributes(path, t,
                     [{"ordinal": 0, "sku": "S1", "values": {"HSN": "54075240"}}])
    assert _cell(path, t, 0, "HSN") == 54075240


def test_write_attributes_leaves_text_columns_as_text(tmp_path):
    """The numeric coercion must apply only to NUMERIC_HEADERS — a Type of '2000'
    would otherwise become a number and break the vocabulary match on reload."""
    t, path = _built(tmp_path)
    write_attributes(path, t,
                     [{"ordinal": 0, "sku": "S1", "values": {"tags": "12345678"}}])
    assert _cell(path, t, 0, "tags") == "12345678"
