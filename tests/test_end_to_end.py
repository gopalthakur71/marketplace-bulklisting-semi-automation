import glob
import io
import os
import warnings

import openpyxl
from PIL import Image

from src.myntra.pipeline import main


def _fake_fetch_factory():
    buf = io.BytesIO()
    Image.new("RGBA", (1000, 1200), (200, 30, 30, 255)).save(buf, "PNG")
    data = buf.getvalue()
    return lambda url: data


def test_full_pipeline(tmp_path):
    warnings.filterwarnings("ignore")
    out_dir = tmp_path / "output"
    result = main(
        template_path="templates/myntra/Myntra-Sku-Template-2026-06-16.xlsx",
        csv_path="tests/fixtures/products_export.csv",
        out_dir=str(out_dir),
        config_dir="config/myntra",
        fetch=_fake_fetch_factory(),
        upload=False,   # don't hit AWS during tests
    )
    assert result["products"] == 2
    assert os.path.exists(result["filled"])
    assert os.path.exists(result["report"])
    wb = openpyxl.load_workbook(result["filled"])
    ws = wb["Sarees"]
    hdr = {ws.cell(3, c).value: c for c in range(1, ws.max_column + 1)}
    assert ws.cell(row=4, column=3).value not in (None, "")
    # Images are stored per-SKU: output/images/<sku>/<n>.jpg
    assert glob.glob(str(out_dir / "images" / "*" / "*.jpg"))
    # Front Image column (74) holds a CDN URL, not a local filename
    front = ws.cell(row=4, column=74).value
    assert front and front.startswith("http")
    # With upload=False the sheet must NOT reference S3 URLs for images it did not
    # upload — it falls back to the source CDN URL instead.
    assert "amazonaws.com" not in front
    assert result["uploaded"] == 0
    # Numeric columns Myntra validates must be stored as numbers, not text.
    for h in ("styleGroupId", "MRP", "ISP", "Year", "Net Quantity"):
        assert ws.cell(row=4, column=hdr[h]).data_type == "n", h
