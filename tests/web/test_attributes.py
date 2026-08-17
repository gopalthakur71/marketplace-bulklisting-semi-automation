import io
import os
import warnings

from fastapi.testclient import TestClient
from PIL import Image

from src.core.models import MappedRow, ImageResult
from src.myntra.fill import fill_template
from src.myntra.template_reader import read_template
from src.web.jobs import store
from src.web.main import create_app
from src.web.settings import Settings
import src.web.routers.attributes as attrs_router
import src.web.routers.generate as gen

V13 = "templates/myntra/Myntra-Sku-Template-2026-07-24.xlsx"


def _client(tmp_path):
    s = Settings(auth_disabled=True, s3_bucket="b",
                 ledger_local_path=str(tmp_path / "led.json"),
                 hsn_local_path=str(tmp_path / "hsn.json"),
                 sku_registry_local_path=str(tmp_path / "reg.json"))
    return TestClient(create_app(s))


def _job(tmp_path, monkeypatch, skus=("S1", "S2"), with_images=True, tags=None,
         hsn=None):
    """A finished job on disk: built workbook + the Shopify export it came from."""
    warnings.filterwarnings("ignore")
    monkeypatch.setattr(gen, "RUNTIME", str(tmp_path / "runtime"))
    job = store.create()
    job_dir = os.path.join(gen.RUNTIME, job.id)
    os.makedirs(job_dir, exist_ok=True)

    t = read_template(V13)
    rows = []
    for s in skus:
        cells = {"vendorSkuCode": s, "brand": "Ijor"}
        if tags is not None:
            cells["tags"] = tags
        if hsn is not None:
            cells["HSN"] = hsn
        rows.append((MappedRow(sku=s, cells=cells), ImageResult(sku=s)))
    xlsx = os.path.join(job_dir, "myntra_filled.xlsx")
    fill_template(V13, t, rows, xlsx)

    with open(os.path.join(job_dir, "products_export.csv"), "w",
              newline="", encoding="utf-8") as fh:
        fh.write("Handle,Title,Variant SKU,Image Src,Image Position\n")
        for i, s in enumerate(skus, start=1):
            img = f"https://cdn.example/{s}.jpg" if with_images else ""
            fh.write(f"h{i},Product {i},{s},{img},1\n")

    job.status = "done"
    job.result = {"filled": xlsx, "report": "", "products": len(skus), "uploaded": 0}
    return job


def test_screen_renders_a_panel_per_sku_with_twelve_selects(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch)
    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    assert r.status_code == 200
    assert r.text.count('class="attr-panel"') == 2
    assert r.text.count('name="attr__0__') == 12          # 12 dropdowns for SKU 1
    assert 'name="sku__0"' in r.text and 'value="S1"' in r.text
    assert "https://cdn.example/S1.jpg" in r.text          # photo from the export
    assert "Product 1" in r.text                           # title from the export


def test_dropdown_options_come_only_from_the_template_vocab(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    assert "<option value=\"\">— choose —</option>" in r.text
    assert ">Banarasi<" in r.text          # a real Type value
    assert ">Zari<" in r.text              # a real Ornamentation value
    assert ">Salmon Pink<" not in r.text   # not in Myntra's colour list


def test_existing_values_are_preselected(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    from src.myntra.attribute_entry import write_attributes
    t = read_template(V13)
    write_attributes(job.result["filled"], t,
                     [{"ordinal": 0, "sku": "S1", "values": {"Border": "Zari"}}])
    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    assert '<option value="Zari" selected>Zari</option>' in r.text
    assert "1/16 filled" in r.text


def test_missing_image_falls_back_to_placeholder(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",), with_images=False)
    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    assert r.status_code == 200
    assert "no photo" in r.text


def test_unknown_job_says_session_expired(tmp_path, monkeypatch):
    monkeypatch.setattr(gen, "RUNTIME", str(tmp_path / "runtime"))
    r = _client(tmp_path).get("/generate/attributes/" + "0" * 32)
    assert r.status_code == 404


def test_live_preview_reconstructs_from_posted_values(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    # column indexes: 0 Prominent Colour .. 5 Type .. 9 Print or Pattern Type
    r = _client(tmp_path).post(f"/generate/attributes/{job.id}/preview", data={
        "ordinal": "0",
        "sku__0": "S1", "attr__0__0": "Blue", "attr__0__3": "Pure Silk",
        "attr__0__5": "Banarasi", "attr__0__9": "Floral"})
    assert r.status_code == 200
    assert "Floral Pure Silk Banarasi Saree" in r.text
    assert "Blue Banarasi sarees" in r.text


def test_live_preview_renders_the_panel_that_asked_not_the_first(tmp_path, monkeypatch):
    """The post carries every included panel's fields; the ordinal decides. Taking
    the first entry rendered panel 0's card into panel 1's slot."""
    job = _job(tmp_path, monkeypatch, skus=("S1", "S2"))
    r = _client(tmp_path).post(f"/generate/attributes/{job.id}/preview", data={
        "ordinal": "1",
        "sku__0": "S1", "attr__0__5": "Banarasi",
        "sku__1": "S2", "attr__1__5": "Chanderi"})
    assert r.status_code == 200
    assert "Chanderi" in r.text
    assert "Banarasi" not in r.text


def test_save_writes_all_skus_and_download_serves_the_updated_file(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1", "S2"))
    client = _client(tmp_path)
    r = client.post(f"/generate/attributes/{job.id}", data={
        "sku__0": "S1", "attr__0__5": "Banarasi", "attr__0__7": "Zari",
        "sku__1": "S2", "attr__1__5": "Chanderi"})
    assert r.status_code == 200
    assert "Saved" in r.text

    t = read_template(V13)
    import openpyxl
    wb = openpyxl.load_workbook(job.result["filled"], data_only=True)
    ws = wb["Sarees"]
    assert ws.cell(row=t.first_data_row,
                   column=t.col_index_by_header["Type"]).value == "Banarasi"
    assert ws.cell(row=t.first_data_row,
                   column=t.col_index_by_header["Border"]).value == "Zari"
    assert ws.cell(row=t.first_data_row + 1,
                   column=t.col_index_by_header["Type"]).value == "Chanderi"
    assert ws.cell(row=t.first_data_row,
                   column=t.col_index_by_header["Pattern"]).value is None
    wb.close()

    d = client.get(f"/generate/download/{job.id}")
    assert d.status_code == 200


def test_save_reopens_with_values_selected(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    client = _client(tmp_path)
    client.post(f"/generate/attributes/{job.id}",
                data={"sku__0": "S1", "attr__0__7": "Zari"})
    r = client.get(f"/generate/attributes/{job.id}")
    assert '<option value="Zari" selected>Zari</option>' in r.text


def test_save_rejects_off_vocab_value_without_writing(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    client = _client(tmp_path)
    r = client.post(f"/generate/attributes/{job.id}",
                    data={"sku__0": "S1", "attr__0__0": "Salmon Pink"})
    assert r.status_code == 200          # htmx-swappable error panel, not a 500
    assert "Prominent Colour" in r.text
    assert "not one of Myntra" in r.text
    t = read_template(V13)
    import openpyxl
    wb = openpyxl.load_workbook(job.result["filled"], data_only=True)
    v = wb["Sarees"].cell(row=t.first_data_row,
                          column=t.col_index_by_header["Prominent Colour"]).value
    wb.close()
    assert v is None


def test_save_keeps_dropdowns_alive_in_the_downloaded_file(tmp_path, monkeypatch):
    """KEY INVARIANT end-to-end: the owner's Excel check must still pass."""
    import openpyxl
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    wb = openpyxl.load_workbook(job.result["filled"])
    before = len(wb["Sarees"].data_validations.dataValidation)
    wb.close()
    _client(tmp_path).post(f"/generate/attributes/{job.id}",
                           data={"sku__0": "S1", "attr__0__7": "Zari"})
    wb = openpyxl.load_workbook(job.result["filled"])
    assert len(wb["Sarees"].data_validations.dataValidation) == before
    wb.close()


def _brand_colour(xlsx, template):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    v = wb["Sarees"].cell(row=template.first_data_row,
                          column=template.col_index_by_header["Brand Colour (Remarks)"]).value
    wb.close()
    return v


def test_save_fills_brand_colour_from_the_prominent_colour(tmp_path, monkeypatch):
    """Myntra rejects a null Brand Colour (Remarks); the app derives it."""
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    _client(tmp_path).post(f"/generate/attributes/{job.id}",
                           data={"sku__0": "S1", "attr__0__0": "Blue"})
    assert _brand_colour(job.result["filled"], read_template(V13)) == "blue"


def test_save_leaves_brand_colour_blank_when_no_colour_chosen(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    _client(tmp_path).post(f"/generate/attributes/{job.id}",
                           data={"sku__0": "S1", "attr__0__7": "Zari"})
    assert _brand_colour(job.result["filled"], read_template(V13)) is None


def test_save_rewrites_brand_colour_when_the_colour_changes(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    client = _client(tmp_path)
    client.post(f"/generate/attributes/{job.id}",
                data={"sku__0": "S1", "attr__0__0": "Blue"})
    client.post(f"/generate/attributes/{job.id}",
                data={"sku__0": "S1", "attr__0__0": "Green"})
    assert _brand_colour(job.result["filled"], read_template(V13)) == "green"


def test_screen_shows_the_derived_brand_colour_read_only(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    client = _client(tmp_path)
    client.post(f"/generate/attributes/{job.id}",
                data={"sku__0": "S1", "attr__0__0": "Blue"})
    r = client.get(f"/generate/attributes/{job.id}")
    assert "Brand Colour (Remarks)" in r.text
    assert "filled automatically" in r.text
    assert ">blue<" in r.text
    # read-only: it must not be a form field the browser posts back
    assert 'name="brand_colour' not in r.text


def test_save_on_expired_job_says_session_expired(tmp_path, monkeypatch):
    monkeypatch.setattr(gen, "RUNTIME", str(tmp_path / "runtime"))
    r = _client(tmp_path).post("/generate/attributes/" + "0" * 32,
                               data={"sku__0": "S1"})
    assert r.status_code == 404


def _cell(xlsx, template, ordinal, header):
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    v = wb["Sarees"].cell(row=template.first_data_row + ordinal,
                          column=template.col_index_by_header[header]).value
    wb.close()
    return v


def test_bulk_save_writes_tags_free_text(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    _client(tmp_path).post(f"/generate/attributes/{job.id}", data={
        "sku__0": "S1", "free__0__0": "saree, cotton, handloom"})
    assert _cell(job.result["filled"], read_template(V13), 0,
                 "tags") == "saree, cotton, handloom"


def test_bulk_save_accepts_tags_that_no_vocabulary_would_allow(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    r = _client(tmp_path).post(f"/generate/attributes/{job.id}", data={
        "sku__0": "S1", "free__0__0": "Salmon Pink"})
    assert "Saved" in r.text
    assert _cell(job.result["filled"], read_template(V13), 0, "tags") == "Salmon Pink"


def test_bulk_save_blank_tags_clears_the_cell(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    client = _client(tmp_path)
    client.post(f"/generate/attributes/{job.id}",
                data={"sku__0": "S1", "free__0__0": "keepme"})
    client.post(f"/generate/attributes/{job.id}",
                data={"sku__0": "S1", "free__0__0": "   "})
    assert _cell(job.result["filled"], read_template(V13), 0, "tags") is None


def test_bulk_save_still_writes_dropdowns_unchanged(tmp_path, monkeypatch):
    """Regression: the extracted helper must not change existing behaviour."""
    job = _job(tmp_path, monkeypatch, skus=("S1", "S2"))
    r = _client(tmp_path).post(f"/generate/attributes/{job.id}", data={
        "sku__0": "S1", "attr__0__5": "Banarasi",
        "sku__1": "S2", "attr__1__5": "Chanderi"})
    assert "Saved" in r.text
    t = read_template(V13)
    assert _cell(job.result["filled"], t, 0, "Type") == "Banarasi"
    assert _cell(job.result["filled"], t, 1, "Type") == "Chanderi"


def _free_input_value(html, ordinal=0, index=0):
    """The value= of one free-text input. A bare `'value=""' in html` check would
    pass on any empty <option>, so match the field itself."""
    import re
    m = re.search(r'name="free__%d__%d"\s+value="([^"]*)"' % (ordinal, index), html)
    assert m, "no free-text input rendered for that ordinal"
    return m.group(1)


def test_panel_renders_a_tags_input_prefilled_from_the_sheet(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",), tags="saree, cotton")
    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    assert _free_input_value(r.text) == "saree, cotton"


# The two name columns. free__N__1 is List View Name, free__N__2 is
# productDisplayName — the order of user_filled_freetext in rules.yaml.

def test_panel_renders_an_input_for_each_name_column(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    assert 'name="free__0__1"' in r.text and "List View Name" in r.text
    assert 'name="free__0__2"' in r.text and "productDisplayName" in r.text


def test_each_free_text_column_carries_its_own_hint(tmp_path, monkeypatch):
    """The shared 'pre-filled from Shopify' hint is a lie for List View Name,
    which the pipeline deliberately leaves blank."""
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    assert "Myntra’s short name for the list view — blank until you write it" in r.text
    assert "the full product name — pre-filled from the Shopify title" in r.text


def test_bulk_save_writes_both_name_columns(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    r = _client(tmp_path).post(f"/generate/attributes/{job.id}", data={
        "sku__0": "S1",
        "free__0__1": "Ijor Cotton Saree",
        "free__0__2": "Ijor Handloom Pure Cotton Saree with Zari Border"})
    assert "Saved" in r.text
    t = read_template(V13)
    assert _cell(job.result["filled"], t, 0, "List View Name") == "Ijor Cotton Saree"
    assert (_cell(job.result["filled"], t, 0, "productDisplayName")
            == "Ijor Handloom Pure Cotton Saree with Zari Border")


def test_a_long_product_display_name_is_not_truncated(tmp_path, monkeypatch):
    """No character cap: the template's instructions sheet states none for these
    two columns, unlike vendorArticleName's 40."""
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    long_name = "Ijor " + "Handwoven " * 12 + "Saree"
    _client(tmp_path).post(f"/generate/attributes/{job.id}", data={
        "sku__0": "S1", "free__0__2": long_name})
    assert _cell(job.result["filled"], read_template(V13), 0,
                 "productDisplayName") == long_name


def test_blank_name_clears_the_cell(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    client = _client(tmp_path)
    client.post(f"/generate/attributes/{job.id}",
                data={"sku__0": "S1", "free__0__2": "Some Name"})
    client.post(f"/generate/attributes/{job.id}",
                data={"sku__0": "S1", "free__0__2": "   "})
    assert _cell(job.result["filled"], read_template(V13), 0,
                 "productDisplayName") is None


def test_both_names_count_toward_the_filled_total_of_sixteen(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    client = _client(tmp_path)
    r = client.post(f"/generate/attributes/{job.id}", data={
        "sku__0": "S1", "free__0__1": "A Name", "free__0__2": "Another Name"})
    assert "Saved" in r.text
    # The bulk response reports how many SKUs were saved; the per-panel counter is
    # what has to reach 2/16, so re-read the screen for it.
    r = client.get(f"/generate/attributes/{job.id}")
    assert "2/16 filled" in r.text


def test_tags_input_is_empty_when_the_sheet_has_none(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    assert _free_input_value(r.text) == ""


def test_filled_count_is_out_of_sixteen_and_counts_tags(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",), tags="festive")
    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    assert "1/16 filled" in r.text


def test_count_span_is_addressable_for_out_of_band_updates(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    assert 'id="attr-count-0"' in r.text


def test_saving_one_panel_writes_only_that_row(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1", "S2"))
    client = _client(tmp_path)
    client.post(f"/generate/attributes/{job.id}",
                data={"sku__0": "S1", "attr__0__5": "Banarasi",
                      "sku__1": "S2", "attr__1__5": "Chanderi"})
    # Now save ONLY panel 0, changing it to a THIRD value. Distinct values per row
    # matter: if row 0's value leaked into row 1 the assertions below must fail.
    r = client.post(f"/generate/attributes/{job.id}/one",
                    data={"ordinal": "0", "sku__0": "S1", "attr__0__5": "Kanjeevaram"})
    assert r.status_code == 200
    t = read_template(V13)
    assert _cell(job.result["filled"], t, 0, "Type") == "Kanjeevaram"
    assert _cell(job.result["filled"], t, 1, "Type") == "Chanderi"  # unchanged


def test_one_panel_save_writes_only_the_requested_ordinal(tmp_path, monkeypatch):
    """The real browser post. Scoping cannot happen in htmx (hx-include only ever
    ADDS fields), so the server must honour the explicit ordinal: a body carrying
    every panel writes one row, and the oob count names THAT panel."""
    job = _job(tmp_path, monkeypatch, skus=("S1", "S2"))
    r = _client(tmp_path).post(f"/generate/attributes/{job.id}/one", data={
        "ordinal": "1",
        "sku__0": "S1", "attr__0__5": "Banarasi", "free__0__0": "row-zero-tags",
        "sku__1": "S2", "attr__1__5": "Chanderi", "free__1__0": "row-one-tags"})
    assert r.status_code == 200
    t = read_template(V13)
    assert _cell(job.result["filled"], t, 1, "Type") == "Chanderi"
    assert _cell(job.result["filled"], t, 1, "tags") == "row-one-tags"
    assert _cell(job.result["filled"], t, 0, "Type") is None       # not written
    assert _cell(job.result["filled"], t, 0, "tags") is None
    assert 'id="attr-count-1"' in r.text
    assert 'id="attr-count-0"' not in r.text
    assert 'hx-swap-oob="true"' in r.text
    assert "2/16 filled" in r.text


def test_one_panel_save_off_vocab_elsewhere_does_not_block_this_panel(tmp_path, monkeypatch):
    """Panel 1's Save must not fail because panel 0 holds a bad value."""
    job = _job(tmp_path, monkeypatch, skus=("S1", "S2"))
    r = _client(tmp_path).post(f"/generate/attributes/{job.id}/one", data={
        "ordinal": "1",
        "sku__0": "S1", "attr__0__0": "Salmon Pink",
        "sku__1": "S2", "attr__1__5": "Chanderi"})
    assert "Saved" in r.text
    assert _cell(job.result["filled"], read_template(V13), 1, "Type") == "Chanderi"


def test_one_panel_save_returns_the_refreshed_count_out_of_band(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    r = _client(tmp_path).post(f"/generate/attributes/{job.id}/one", data={
        "ordinal": "0",
        "sku__0": "S1", "attr__0__5": "Banarasi", "free__0__0": "festive"})
    assert 'id="attr-count-0"' in r.text
    assert 'hx-swap-oob="true"' in r.text
    assert "2/16 filled" in r.text
    assert "Saved" in r.text


def test_one_panel_save_rejects_off_vocab_inline_without_writing(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    r = _client(tmp_path).post(
        f"/generate/attributes/{job.id}/one",
        data={"ordinal": "0", "sku__0": "S1", "attr__0__0": "Salmon Pink"})
    assert r.status_code == 200                 # inline error, not a 500
    assert "not one of Myntra" in r.text
    assert 'hx-swap-oob' not in r.text          # count must NOT be updated
    assert _cell(job.result["filled"], read_template(V13), 0,
                 "Prominent Colour") is None


def test_one_panel_save_keeps_dropdowns_alive(tmp_path, monkeypatch):
    import openpyxl
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    wb = openpyxl.load_workbook(job.result["filled"])
    before = len(wb["Sarees"].data_validations.dataValidation)
    wb.close()
    _client(tmp_path).post(f"/generate/attributes/{job.id}/one",
                           data={"ordinal": "0", "sku__0": "S1",
                                 "attr__0__7": "Zari"})
    wb = openpyxl.load_workbook(job.result["filled"])
    assert len(wb["Sarees"].data_validations.dataValidation) == before
    wb.close()


def test_one_panel_save_holds_the_write_lock(tmp_path, monkeypatch):
    """The lock must be held across the read-modify-write, not merely to exist."""
    import src.web.routers.attributes as attrs
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    seen = {}
    real = attrs.write_attributes

    def spy(*a, **k):
        seen["locked"] = attrs._WRITE_LOCK.locked()
        return real(*a, **k)

    monkeypatch.setattr(attrs, "write_attributes", spy)
    _client(tmp_path).post(f"/generate/attributes/{job.id}/one",
                           data={"ordinal": "0", "sku__0": "S1",
                                 "attr__0__7": "Zari"})
    assert seen["locked"] is True


def test_one_panel_save_on_expired_job_says_session_expired(tmp_path, monkeypatch):
    monkeypatch.setattr(gen, "RUNTIME", str(tmp_path / "runtime"))
    r = _client(tmp_path).post("/generate/attributes/" + "0" * 32 + "/one",
                               data={"sku__0": "S1"})
    assert r.status_code == 404


def test_panel_save_button_declares_its_own_ordinal(tmp_path, monkeypatch):
    """Markup half of the guarantee — the behavioural half is
    test_one_panel_save_writes_only_the_requested_ordinal. Each button must state
    WHICH panel it is; the server cannot infer it from the posted keys."""
    import re
    job = _job(tmp_path, monkeypatch, skus=("S1", "S2"))
    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    buttons = re.findall(r"<button[^>]*hx-post=\"/generate/attributes/[^\"]+/one\"[^>]*>",
                         r.text)
    assert len(buttons) == 2
    # Load-bearing: a submit button would post the page rather than fire htmx.
    assert all('type="button"' in b for b in buttons)
    assert '"ordinal": 0' in buttons[0] and '"ordinal": 1' in buttons[1]
    assert 'id="attr-save-0"' in r.text and 'id="attr-save-1"' in r.text


def test_panels_are_not_wrapped_in_a_form(tmp_path, monkeypatch):
    """htmx posts an enclosing form's EVERY field on any non-GET and hx-include can
    only add to that set. A <form> here would silently un-scope every panel."""
    job = _job(tmp_path, monkeypatch, skus=("S1", "S2"))
    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    assert "<form" not in r.text
    assert 'id="attr-form"' in r.text
    assert 'hx-include="#attr-form"' in r.text     # bulk save still posts them all


def test_live_preview_element_declares_its_own_ordinal(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1", "S2"))
    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    assert r.text.count("""hx-vals='{"ordinal": 1}'""") == 2   # grid + save button


def test_one_panel_save_with_no_parseable_entries_touches_no_panel(tmp_path, monkeypatch):
    """A post with no sku__N / attr__N__* keys must not fall back to ordinal 0 —
    that would silently stamp a wrong out-of-band count onto panel 0, an untouched
    panel, while whichever panel the user actually clicked does nothing."""
    job = _job(tmp_path, monkeypatch, skus=("S1", "S2"))
    r = _client(tmp_path).post(f"/generate/attributes/{job.id}/one",
                               data={"unrelated_field": "x"})
    assert r.status_code == 200
    assert 'hx-swap-oob' not in r.text
    assert "Saved" not in r.text


def test_one_panel_save_for_an_absent_ordinal_touches_no_panel(tmp_path, monkeypatch):
    """The ordinal is explicit, but that panel posted no fields of its own."""
    job = _job(tmp_path, monkeypatch, skus=("S1", "S2"))
    r = _client(tmp_path).post(f"/generate/attributes/{job.id}/one",
                               data={"ordinal": "1", "sku__0": "S1",
                                     "attr__0__5": "Banarasi"})
    assert r.status_code == 200
    assert 'hx-swap-oob' not in r.text
    assert "Saved" not in r.text
    assert _cell(job.result["filled"], read_template(V13), 0, "Type") is None


def test_untouched_tags_field_round_trips_unchanged(tmp_path, monkeypatch):
    """Spec test 6. The box is pre-filled from the sheet; a user who never touches
    it still posts it back, and that must not blank or alter the cell."""
    job = _job(tmp_path, monkeypatch, skus=("S1",), tags="saree, cotton, handloom")
    client = _client(tmp_path)
    rendered = _free_input_value(client.get(f"/generate/attributes/{job.id}").text)
    assert rendered == "saree, cotton, handloom"
    # Post back exactly what the browser holds, changing only a dropdown.
    client.post(f"/generate/attributes/{job.id}/one",
                data={"ordinal": "0", "sku__0": "S1", "attr__0__5": "Banarasi",
                      "free__0__0": rendered})
    t = read_template(V13)
    assert _cell(job.result["filled"], t, 0, "tags") == "saree, cotton, handloom"
    assert _cell(job.result["filled"], t, 0, "Type") == "Banarasi"
    # and it comes back pre-filled again on reload
    assert _free_input_value(
        client.get(f"/generate/attributes/{job.id}").text) == "saree, cotton, handloom"


def test_hsn_renders_prefilled_from_the_sheet(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",), hsn="54075240")
    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    assert 'name="hsn__0"' in r.text
    assert 'value="54075240"' in r.text


def test_hsn_gap_banner_counts_missing_codes(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1", "S2"))     # neither has an HSN
    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    assert "2 SKUs still need an HSN" in r.text


def test_saving_a_valid_hsn_writes_it_and_clears_the_banner(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    client = _client(tmp_path)
    r = client.post(f"/generate/attributes/{job.id}/one",
                    data={"ordinal": 0, "sku__0": "S1", "hsn__0": "54075240"})
    assert r.status_code == 200
    assert "Saved" in r.text
    assert "Every SKU has an HSN" in r.text          # out-of-band banner refresh
    assert _cell(job.result["filled"], read_template(V13), 0, "HSN") == 54075240


def test_hsn_counts_toward_the_filled_total_of_sixteen(tmp_path, monkeypatch):
    # 12 dropdowns + tags + HSN. _filled_count is the single shared definition,
    # so the panel header and the save result cannot disagree.
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    r = _client(tmp_path).post(f"/generate/attributes/{job.id}/one",
                               data={"ordinal": 0, "sku__0": "S1",
                                     "hsn__0": "54075240"})
    assert "1/16 filled" in r.text


def test_saving_a_bad_hsn_is_rejected_and_writes_nothing(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",), hsn="54075240")
    client = _client(tmp_path)
    r = client.post(f"/generate/attributes/{job.id}/one",
                    data={"ordinal": 0, "sku__0": "S1", "hsn__0": "5407"})
    assert "Not saved" in r.text and "8-digit" in r.text
    assert 'hx-swap-oob' not in r.text               # no count or banner disturbed
    assert _cell(job.result["filled"], read_template(V13), 0,
                 "HSN") == 54075240                  # unchanged


def test_a_panel_that_posts_no_hsn_field_leaves_the_cell_alone(tmp_path, monkeypatch):
    """Only a posted hsn__N clears the cell. A save that never carried the field —
    e.g. the live-preview path or an older cached page — must not blank it."""
    job = _job(tmp_path, monkeypatch, skus=("S1",), hsn="54075240")
    _client(tmp_path).post(f"/generate/attributes/{job.id}/one",
                           data={"ordinal": 0, "sku__0": "S1",
                                 "attr__0__5": "Banarasi"})
    assert _cell(job.result["filled"], read_template(V13), 0, "HSN") == 54075240


def test_blank_hsn_clears_the_cell(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",), hsn="54075240")
    r = _client(tmp_path).post(f"/generate/attributes/{job.id}/one",
                               data={"ordinal": 0, "sku__0": "S1", "hsn__0": "  "})
    assert "Saved" in r.text
    assert "1 SKU still needs an HSN" in r.text      # banner counts it again
    assert _cell(job.result["filled"], read_template(V13), 0, "HSN") is None


def test_screen_reads_the_workbook_only_once(tmp_path, monkeypatch):
    """The gap count is derived from the rows the panels already read. A second
    read_filled_rows here reloads the whole workbook — the slowest single thing
    this screen does — to recount something already in hand."""
    import src.web.routers.attributes as attrs
    job = _job(tmp_path, monkeypatch, skus=("S1", "S2"))
    calls = []
    real = attrs.read_filled_rows

    def spy(*a, **k):
        calls.append(1)
        return real(*a, **k)

    monkeypatch.setattr(attrs, "read_filled_rows", spy)
    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    assert r.status_code == 200
    assert "2 SKUs still need an HSN" in r.text      # the count is still right
    assert len(calls) == 1


def test_bulk_save_refreshes_the_banner_at_the_top_level_of_the_fragment(tmp_path,
                                                                        monkeypatch):
    """htmx processes hx-swap-oob only on TOP-LEVEL elements of a response
    fragment — nested inside the result panel it is silently ignored and the
    banner never refreshes. A plain "the span is present" assertion passes either
    way, so this pins the position: it must follow the panel div, not sit in it."""
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    r = _client(tmp_path).post(f"/generate/attributes/{job.id}",
                               data={"sku__0": "S1", "hsn__0": "54075240"})
    assert "Saved" in r.text
    assert 'id="hsn-gap-banner"' in r.text
    assert r.text.index('id="hsn-gap-banner"') > r.text.rindex("</div>")
    assert "Every SKU has an HSN" in r.text


def test_bulk_save_rejected_emits_no_banner(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    r = _client(tmp_path).post(f"/generate/attributes/{job.id}",
                               data={"sku__0": "S1", "hsn__0": "5407"})
    assert "Nothing was saved" in r.text
    assert 'hx-swap-oob' not in r.text


def test_per_panel_save_writes_only_the_requested_panels_hsn(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1", "S2"))
    client = _client(tmp_path)
    client.post(f"/generate/attributes/{job.id}/one",
                data={"ordinal": 1, "sku__0": "S1", "hsn__0": "11111111",
                      "sku__1": "S2", "hsn__1": "22222222"})
    t = read_template(V13)
    assert _cell(job.result["filled"], t, 0, "HSN") is None      # panel 0 untouched
    assert _cell(job.result["filled"], t, 1, "HSN") == 22222222


def test_saving_an_hsn_corrects_the_sku_registry(tmp_path, monkeypatch):
    """The fix-flow rebuild pins HSN from the registry, so a correction made here
    has to reach it or a later rebuild restores the stale build-time code."""
    from src.myntra.sku_registry import read_registry, record
    from src.web.settings import sku_registry_store
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    s = Settings(auth_disabled=True, s3_bucket="b",
                 ledger_local_path=str(tmp_path / "led.json"),
                 hsn_local_path=str(tmp_path / "hsn.json"),
                 sku_registry_local_path=str(tmp_path / "reg.json"))
    record(sku_registry_store(s), "S1", "hash-1", 7, "50072010")

    TestClient(create_app(s)).post(
        f"/generate/attributes/{job.id}/one",
        data={"ordinal": 0, "sku__0": "S1", "hsn__0": "54075240"})

    entry = read_registry(sku_registry_store(s))["S1"]
    assert entry["hsn"] == "54075240"
    assert entry["content_hash"] == "hash-1"        # nothing else disturbed


def test_a_rejected_save_does_not_move_the_registry(tmp_path, monkeypatch):
    """The registry update sits inside the same try as the write: a save that
    never reached the sheet must not leave the registry claiming it did."""
    from src.myntra.sku_registry import read_registry, record
    from src.web.settings import sku_registry_store
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    s = Settings(auth_disabled=True, s3_bucket="b",
                 ledger_local_path=str(tmp_path / "led.json"),
                 hsn_local_path=str(tmp_path / "hsn.json"),
                 sku_registry_local_path=str(tmp_path / "reg.json"))
    record(sku_registry_store(s), "S1", "hash-1", 7, "50072010")

    TestClient(create_app(s)).post(
        f"/generate/attributes/{job.id}/one",
        data={"ordinal": 0, "sku__0": "S1", "hsn__0": "5407"})

    assert read_registry(sku_registry_store(s))["S1"]["hsn"] == "50072010"


def _registry(tmp_path):
    import json
    with open(tmp_path / "reg.json", encoding="utf-8") as fh:
        return json.load(fh)


def test_saving_a_name_pins_it_in_the_sku_registry(tmp_path, monkeypatch):
    """The attribute screen is the only place these names are authored, so the
    pin has to be written here or a later rebuild reverts to the Shopify title."""
    from src.myntra.sku_registry import record
    from src.web.settings import Settings, sku_registry_store
    s = Settings(auth_disabled=True, s3_bucket="b",
                 ledger_local_path=str(tmp_path / "led.json"),
                 hsn_local_path=str(tmp_path / "hsn.json"),
                 sku_registry_local_path=str(tmp_path / "reg.json"))
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    record(sku_registry_store(s), "S1", "hash-1", 42, "50072010")

    TestClient(create_app(s)).post(f"/generate/attributes/{job.id}", data={
        "sku__0": "S1",
        "free__0__1": "Ijor Saree",
        "free__0__2": "Ijor Handloom Pure Cotton Saree"})

    entry = _registry(tmp_path)["S1"]
    assert entry["names"] == {"List View Name": "Ijor Saree",
                              "productDisplayName": "Ijor Handloom Pure Cotton Saree"}
    assert entry["style_group_id"] == 42          # untouched
    assert entry["hsn"] == "50072010"             # untouched


def test_a_rejected_save_pins_no_name(tmp_path, monkeypatch):
    """The registry must not move when the write was refused — same rule the HSN
    pin follows."""
    from src.myntra.sku_registry import record
    from src.web.settings import Settings, sku_registry_store
    s = Settings(auth_disabled=True, s3_bucket="b",
                 ledger_local_path=str(tmp_path / "led.json"),
                 hsn_local_path=str(tmp_path / "hsn.json"),
                 sku_registry_local_path=str(tmp_path / "reg.json"))
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    record(sku_registry_store(s), "S1", "hash-1", 42, "50072010")

    r = TestClient(create_app(s)).post(f"/generate/attributes/{job.id}", data={
        "sku__0": "S1",
        "attr__0__0": "Salmon Pink",              # off-vocabulary -> whole save fails
        "free__0__2": "Ijor Handloom Pure Cotton Saree"})
    assert "not one of Myntra" in r.text
    assert not _registry(tmp_path)["S1"].get("names")


def test_panel_photo_falls_back_to_the_sheet_front_image(tmp_path, monkeypatch):
    """An adopted upload has no Shopify export, so the only photo available is the
    URL already in the sheet. Without this every uploaded sheet shows 'no photo'."""
    warnings.filterwarnings("ignore")
    monkeypatch.setattr(gen, "RUNTIME", str(tmp_path / "runtime"))
    job = store.create()
    job_dir = os.path.join(gen.RUNTIME, job.id)
    os.makedirs(job_dir, exist_ok=True)
    t = read_template(V13)
    row = MappedRow(sku="S1", cells={
        "vendorSkuCode": "S1", "brand": "Ijor",
        "Front Image": "https://cdn.example/S1-front.jpg"})
    xlsx = os.path.join(job_dir, "myntra_filled.xlsx")
    fill_template(V13, t, [(row, ImageResult(sku="S1"))], xlsx)
    job.status = "done"
    job.result = {"filled": xlsx, "origin": "upload", "filename": "s.xlsx"}

    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    assert r.status_code == 200
    assert "https://cdn.example/S1-front.jpg" in r.text


def test_panel_photo_ignores_a_non_url_front_image(tmp_path, monkeypatch):
    """fill.py falls back to bare local filenames when S3 is off. Rendering one as
    an <img src> would show a broken image; the placeholder is honest."""
    warnings.filterwarnings("ignore")
    monkeypatch.setattr(gen, "RUNTIME", str(tmp_path / "runtime"))
    job = store.create()
    job_dir = os.path.join(gen.RUNTIME, job.id)
    os.makedirs(job_dir, exist_ok=True)
    t = read_template(V13)
    row = MappedRow(sku="S1", cells={
        "vendorSkuCode": "S1", "brand": "Ijor", "Front Image": "1.jpg"})
    xlsx = os.path.join(job_dir, "myntra_filled.xlsx")
    fill_template(V13, t, [(row, ImageResult(sku="S1"))], xlsx)
    job.status = "done"
    job.result = {"filled": xlsx, "origin": "upload", "filename": "s.xlsx"}

    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    assert "no photo" in r.text


def test_panel_offers_a_file_input_for_every_myntra_image_slot(tmp_path, monkeypatch):
    """Myntra rejects one specific shot ('front image is pixelated'), so each slot
    needs its own picker — replacing all seven is the same path, just more files."""
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    r = _client(tmp_path).get(f"/generate/attributes/{job.id}")
    assert r.status_code == 200
    for slot in range(1, 8):
        assert f'name="img__0__{slot}"' in r.text
    assert "Front Image" in r.text
    assert "Additional Image 2" in r.text


def _png_bytes(w, h):
    buf = io.BytesIO()
    Image.new("RGB", (w, h), "blue").save(buf, "PNG")
    return buf.getvalue()


def test_replacing_an_image_writes_its_url_into_the_sheet(tmp_path, monkeypatch):
    """The point of the feature: a new photo must reach the workbook as a URL
    Myntra can fetch."""
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    monkeypatch.setattr(attrs_router, "load_specs", lambda: {
        "min_width": 700, "min_height": 700, "max_bytes": 10485760, "quality": 90,
        "public_base_url": "https://cdn.example/myntra", "s3_bucket": "b",
        "s3_prefix": "myntra", "s3_upload": True})
    monkeypatch.setattr(attrs_router, "host",
                        lambda prepared, specs, out_dir: [
                            f"https://cdn.example/myntra/{k}" for _, k in prepared])

    r = _client(tmp_path).post(
        f"/generate/attributes/{job.id}/images",
        data={"ordinal": "0", "sku__0": "S1"},
        files={"img__0__1": ("new.png", _png_bytes(800, 800), "image/png")})
    assert r.status_code == 200
    assert "Front Image" in r.text

    from src.myntra.preview import read_filled_rows
    from src.myntra.template_reader import read_template
    rows = read_filled_rows(job.result["filled"], read_template(V13))
    assert rows[0]["Front Image"].startswith("https://cdn.example/myntra/S1/1-")


def test_an_undersized_replacement_fails_only_its_own_slot(tmp_path, monkeypatch):
    """A bad photo in one slot must not discard the good photo supplied alongside it."""
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    monkeypatch.setattr(attrs_router, "load_specs", lambda: {
        "min_width": 700, "min_height": 700, "max_bytes": 10485760, "quality": 90,
        "public_base_url": "https://cdn.example/myntra", "s3_bucket": "b",
        "s3_prefix": "myntra", "s3_upload": True})
    monkeypatch.setattr(attrs_router, "host",
                        lambda prepared, specs, out_dir: [
                            f"https://cdn.example/myntra/{k}" for _, k in prepared])

    r = _client(tmp_path).post(
        f"/generate/attributes/{job.id}/images",
        data={"ordinal": "0", "sku__0": "S1"},
        files={"img__0__1": ("small.png", _png_bytes(300, 300), "image/png"),
               "img__0__2": ("good.png", _png_bytes(800, 800), "image/png")})
    assert "300x300" in r.text
    from src.myntra.preview import read_filled_rows
    from src.myntra.template_reader import read_template
    rows = read_filled_rows(job.result["filled"], read_template(V13))
    assert rows[0]["Side Image"].startswith("https://cdn.example/myntra/S1/2-")


def test_unconfigured_hosting_reports_instead_of_writing_a_local_path(tmp_path, monkeypatch):
    job = _job(tmp_path, monkeypatch, skus=("S1",))
    monkeypatch.setattr(attrs_router, "load_specs", lambda: {
        "min_width": 700, "min_height": 700, "max_bytes": 10485760, "quality": 90,
        "public_base_url": "", "s3_bucket": "", "s3_upload": False})
    r = _client(tmp_path).post(
        f"/generate/attributes/{job.id}/images",
        data={"ordinal": "0", "sku__0": "S1"},
        files={"img__0__1": ("new.png", _png_bytes(800, 800), "image/png")})
    assert r.status_code == 200
    assert "not configured" in r.text
