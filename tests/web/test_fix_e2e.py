import io
import os
import openpyxl
from PIL import Image
from fastapi.testclient import TestClient

from src.web.main import create_app
from src.web.settings import Settings
import src.web.routers.fix as fixmod


def _fake_image_bytes():
    buf = io.BytesIO()
    Image.new("RGBA", (1000, 1200), (200, 30, 30, 255)).save(buf, "PNG")
    return buf.getvalue()


def _client(tmp_path):
    return TestClient(create_app(Settings(
        auth_disabled=True, s3_bucket="b",
        explanation_store_path=str(tmp_path / "expl.json"),
        correction_log_path=str(tmp_path / "log.json"))))


def _sku_xlsx_bytes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sarees"
    headers = ["STATUS", "SYSTEM ERROR MESSAGE", "vendorSkuCode", "brand",
               "Manufacturer Name and Address with Pincode",
               "Packer Name and Address with Pincode", "Front Image"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    # AAA: address auto-fix (correctable); IMG: flat-shot image (explain-only)
    ws.append([]) if False else None
    ws.cell(row=4, column=1, value="SKU_VALIDATION_FAILED")
    ws.cell(row=4, column=2, value="Manufacturer and packer information is incomplete")
    ws.cell(row=4, column=3, value="AAA")
    ws.cell(row=5, column=1, value="SKU_VALIDATION_FAILED")
    ws.cell(row=5, column=2, value="Primary image appears to be a flat shot")
    ws.cell(row=5, column=3, value="IMG")
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_surface_a_end_to_end_excludes_explain_only(tmp_path, monkeypatch):
    # Use the real template + constants; keep fill_template light by pointing at
    # the repo template. No network: images come from cells, none present here.
    client = _client(tmp_path)
    up = client.post("/fix", files={"file": ("wLf4susb_file.xlsx", _sku_xlsx_bytes(),
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert up.status_code == 200
    assert "AAA" in up.text            # correctable group
    assert "flat shot" in up.text      # explain-only group
    fix_id = up.headers["x-fix-id"]

    r = client.post(f"/fix/apply/{fix_id}", data={})
    assert r.status_code == 200
    assert "Download corrected xlsx" in r.text
    assert "IMG" in r.text             # surfaced as manual-needed, not in the file

    # correction log recorded the AAA address fix
    from src.myntra.correction_log import read_log
    from src.web.settings import LocalJsonStore
    log = read_log(LocalJsonStore(str(tmp_path / "log.json")))
    assert any(rec["sku"] == "AAA" for rec in log)


def test_surface_b_end_to_end(monkeypatch, tmp_path):
    client = _client(tmp_path)
    captured = {}

    def fake_regen(skus, settings, out_dir, csv_path=None):
        captured["csv_path"] = csv_path
        with open(csv_path, "rb") as fh:
            captured["bytes"] = fh.read()
        path = f"{out_dir}/myntra_filled.xlsx"
        with open(path, "wb") as fh:
            fh.write(b"rebuilt")
        return {"written": 1, "file": path, "fixed": list(skus or []),
                "could_not_rebuild": [], "manual_needed": [], "dropped": [],
                "rejected": {}, "changed": {}}

    monkeypatch.setattr(fixmod, "regenerate_surface_b", fake_regen)

    # "manufacturer and packer information is incomplete" is a real configured
    # rule (auto_fix/address) so this row is genuinely correctable -> non-empty
    # skus set for Surface B. A reason that matches no rule would be explain_only
    # and correctly produce an empty rebuild set (see FIX 1 / test_fix.py).
    listings = (b'"style status","seller sku code","onhold reason","style id"\r\n'
                b'"PMR","127SDE826NSB","manufacturer and packer information is incomplete","43214808"\r\n')
    up = client.post("/fix", files={"file": ("MDirect_Listings_Report.csv", listings, "text/csv")})
    assert up.status_code == 200
    fix_id = up.headers["x-fix-id"]
    # Surface B now requires the Shopify export; attach it so the rebuild proceeds.
    r = client.post(f"/fix/apply/{fix_id}", files={
        "products_export": ("products_export.csv", b"Handle,Variant SKU\nx,127SDE826NSB\n", "text/csv")})
    assert r.status_code == 200
    assert "Download corrected xlsx" in r.text
    # the uploaded export was threaded through to the rebuild as a real file
    assert captured["csv_path"] and os.path.exists(captured["csv_path"])
    assert captured["bytes"] == b"Handle,Variant SKU\nx,127SDE826NSB\n"


def test_surface_b_real_rebuild_end_to_end(monkeypatch, tmp_path):
    """No monkeypatch of regenerate_surface_b: drive the REAL pipeline from an
    uploaded products export. This is the path the prod bug crashed on (the
    pipeline defaulted csv_path to a missing input/ file); only the image fetch
    and S3 upload are stubbed so the test stays offline."""
    import src.myntra.pipeline as pipe
    import src.myntra.corrector as corrector
    import src.core.s3_upload as s3
    from src.core.images import process_images as real_process_images

    img = _fake_image_bytes()
    monkeypatch.setattr(pipe, "process_images",
                        lambda p, specs, out_dir: real_process_images(
                            p, specs, out_dir, fetch=lambda url: img))
    monkeypatch.setattr(s3, "upload_images", lambda *a, **k: [])
    # Unknown registry -> fresh sequential ids, no pins needed for the rebuild.
    monkeypatch.setattr(corrector, "read_registry", lambda store: {})
    monkeypatch.setattr(corrector, "sku_registry_store", lambda s: object())

    client = _client(tmp_path)
    listings = (b'"style status","seller sku code","onhold reason","style id"\r\n'
                b'"PMR","TST001","manufacturer and packer information is incomplete","43214808"\r\n')
    up = client.post("/fix", files={"file": ("MDirect_Listings_Report.csv", listings, "text/csv")})
    assert up.status_code == 200
    fix_id = up.headers["x-fix-id"]

    with open("tests/fixtures/products_export.csv", "rb") as fh:
        export_bytes = fh.read()
    r = client.post(f"/fix/apply/{fix_id}", files={
        "products_export": ("products_export.csv", export_bytes, "text/csv")})
    assert r.status_code == 200
    assert "Download corrected xlsx" in r.text

    dl = client.get(f"/fix/download/{fix_id}")
    assert dl.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(dl.content))
    ws = wb["Sarees"]
    assert ws.cell(row=4, column=3).value not in (None, "")  # TST001 rebuilt into the sheet


def test_manual_rebuild_real_pipeline_end_to_end(monkeypatch, tmp_path):
    """action=manual drives the REAL pipeline from an uploaded export for an
    explain-only SKU and downloads a valid xlsx. Only image fetch + S3 are stubbed."""
    import src.myntra.pipeline as pipe
    import src.myntra.corrector as corrector
    import src.core.s3_upload as s3
    from src.core.images import process_images as real_process_images

    img = _fake_image_bytes()
    monkeypatch.setattr(pipe, "process_images",
                        lambda p, specs, out_dir: real_process_images(
                            p, specs, out_dir, fetch=lambda url: img))
    monkeypatch.setattr(s3, "upload_images", lambda *a, **k: [])
    monkeypatch.setattr(corrector, "read_registry", lambda store: {})
    monkeypatch.setattr(corrector, "sku_registry_store", lambda s: object())

    client = _client(tmp_path)
    # A reason that matches NO configured rule -> explain_only (plain) for listings_report.
    listings = (b'"style status","seller sku code","onhold reason","style id"\r\n'
                b'"PMR","TST001","image resolution is too low, no rule matches this wording","43214808"\r\n')
    up = client.post("/fix", files={"file": ("MDirect_Listings_Report.csv", listings, "text/csv")})
    assert up.status_code == 200
    assert "Download listing file" in up.text          # manual button rendered
    fix_id = up.headers["x-fix-id"]

    with open("tests/fixtures/products_export.csv", "rb") as fh:
        export_bytes = fh.read()
    r = client.post(f"/fix/apply/{fix_id}",
                    data={"action": "manual"},
                    files={"products_export": ("products_export.csv", export_bytes, "text/csv")})
    assert r.status_code == 200
    assert "Download corrected xlsx" in r.text

    dl = client.get(f"/fix/download/{fix_id}")
    assert dl.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(dl.content))
    ws = wb["Sarees"]
    assert ws.cell(row=4, column=3).value not in (None, "")   # TST001 rebuilt into the sheet
