# tests/test_corrector.py
import openpyxl
from src.myntra.error_reader import load_rules, read_errors
from src.myntra.template_reader import read_template
from src.myntra.corrector import plan_corrections, correct

TEMPLATE = "templates/myntra/Myntra-Sku-Template-2026-06-16.xlsx"
IMG = "https://ijorethnicpartners.s3.ap-south-1.amazonaws.com/myntra"
# correct(row_errors, template, template_path, constants, answers, drops, out_path)


def _make_resub(path, rows):
    """rows = list of dict(status, message, cells={header: value})."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sarees"
    headers = ["STATUS", "SYSTEM ERROR MESSAGE", "styleGroupId", "vendorSkuCode",
               "brand", "MRP", "ISP", "Prominent Colour", "Brand Colour (Remarks)",
               "Front Image"]
    for c, h in enumerate(headers, start=1):
        ws.cell(row=3, column=c, value=h)
    r = 4
    for row in rows:
        ws.cell(row=r, column=1, value=row["status"])
        ws.cell(row=r, column=2, value=row["message"])
        for c, h in enumerate(headers, start=1):
            if h in ("STATUS", "SYSTEM ERROR MESSAGE"):
                continue
            ws.cell(row=r, column=c, value=row["cells"].get(h))
        r += 1
    wb.save(path)


def test_plan_marks_drop_and_manual(tmp_path):
    p = tmp_path / "resub.xlsx"
    _make_resub(p, [
        {"status": "SKU_VALIDATION_FAILED",
         "message": "Brand Colour (Remarks) cannot be null",
         "cells": {"styleGroupId": "11", "vendorSkuCode": "78SAZ125BSI",
                   "brand": "Ijor Ethnic Partners", "Front Image": f"{IMG}/78SAZ125BSI/1.jpg"}},
        {"status": "SKU_VALIDATION_FAILED",
         "message": "Seller Sku Code 165SDE226RSG is already registered",
         "cells": {"styleGroupId": "12", "vendorSkuCode": "165SDE226RSG",
                   "brand": "Ijor Ethnic Partners", "Front Image": f"{IMG}/165SDE226RSG/1.jpg"}},
    ])
    rules = load_rules()
    errs = read_errors(str(p), rules)
    plan = plan_corrections(errs)
    assert plan["drop"] == ["165SDE226RSG"]
    assert plan["manual"][0]["sku"] == "78SAZ125BSI"
    assert plan["manual"][0]["field"] == "Prominent Colour"


def test_correct_drops_and_applies_answer(tmp_path):
    p = tmp_path / "resub.xlsx"
    _make_resub(p, [
        {"status": "SKU_VALIDATION_FAILED",
         "message": "Brand Colour (Remarks) cannot be null",
         "cells": {"styleGroupId": "11", "vendorSkuCode": "78SAZ125BSI",
                   "brand": "Ijor Ethnic Partners", "Front Image": f"{IMG}/78SAZ125BSI/1.jpg"}},
        {"status": "SKU_VALIDATION_FAILED",
         "message": "Seller Sku Code 165SDE226RSG is already registered",
         "cells": {"styleGroupId": "12", "vendorSkuCode": "165SDE226RSG",
                   "brand": "Ijor Ethnic Partners", "Front Image": f"{IMG}/165SDE226RSG/1.jpg"}},
    ])
    rules = load_rules()
    errs = read_errors(str(p), rules)
    template = read_template(TEMPLATE)
    out = tmp_path / "corrected.xlsx"
    summary = correct(
        errs, template, TEMPLATE, constants={},
        answers={"78SAZ125BSI": {"Prominent Colour": "White"}},
        drops={"165SDE226RSG"}, out_path=str(out),
    )
    assert summary["written"] == 1
    assert summary["dropped"] == ["165SDE226RSG"]
    ws = openpyxl.load_workbook(out)["Sarees"]
    hdr = {ws.cell(3, c).value: c for c in range(1, ws.max_column + 1)}
    # only the kept SKU is written, with the chosen colour and its image URL
    assert ws.cell(4, hdr["vendorSkuCode"]).value == "78SAZ125BSI"
    assert ws.cell(4, hdr["Prominent Colour"]).value == "White"
    assert ws.cell(4, hdr["Front Image"]).value == f"{IMG}/78SAZ125BSI/1.jpg"
    assert ws.cell(5, hdr["vendorSkuCode"]).value in (None, "")  # dropped SKU not written


def test_correct_backfills_empty_isp_from_mrp(tmp_path):
    p = tmp_path / "resub.xlsx"
    _make_resub(p, [
        {"status": "SKU_VALIDATION_FAILED",
         "message": "ISP cannot be empty for DIY source",
         "cells": {"styleGroupId": "11", "vendorSkuCode": "ABC123",
                   "brand": "Ijor Ethnic Partners", "MRP": "2999", "ISP": None,
                   "Front Image": f"{IMG}/ABC123/1.jpg"}},
    ])
    rules = load_rules()
    errs = read_errors(str(p), rules)
    template = read_template(TEMPLATE)
    out = tmp_path / "corrected.xlsx"
    summary = correct(errs, template, TEMPLATE, constants={}, answers={},
                      drops=set(), out_path=str(out))
    ws = openpyxl.load_workbook(out)["Sarees"]
    hdr = {ws.cell(3, c).value: c for c in range(1, ws.max_column + 1)}
    # empty ISP backfilled from MRP, written as a real number (fill coerces)
    assert ws.cell(4, hdr["ISP"]).value == 2999
    assert "ISP" in summary["changed"]["ABC123"]


def test_correct_validates_colour_answer(tmp_path):
    p = tmp_path / "resub.xlsx"
    _make_resub(p, [
        {"status": "SKU_VALIDATION_FAILED",
         "message": "Brand Colour (Remarks) cannot be null",
         "cells": {"styleGroupId": "11", "vendorSkuCode": "AAA",
                   "brand": "Ijor Ethnic Partners", "Front Image": f"{IMG}/AAA/1.jpg"}},
        {"status": "SKU_VALIDATION_FAILED",
         "message": "Brand Colour (Remarks) cannot be null",
         "cells": {"styleGroupId": "12", "vendorSkuCode": "BBB",
                   "brand": "Ijor Ethnic Partners", "Front Image": f"{IMG}/BBB/1.jpg"}},
    ])
    rules = load_rules()
    errs = read_errors(str(p), rules)
    template = read_template(TEMPLATE)
    out = tmp_path / "corrected.xlsx"
    summary = correct(
        errs, template, TEMPLATE, constants={},
        answers={"AAA": {"Prominent Colour": "white"},     # valid, wrong case
                 "BBB": {"Prominent Colour": "Nosuchclr"}},  # not a dropdown value
        drops=set(), out_path=str(out),
    )
    ws = openpyxl.load_workbook(out)["Sarees"]
    hdr = {ws.cell(3, c).value: c for c in range(1, ws.max_column + 1)}
    rows = {ws.cell(r, hdr["vendorSkuCode"]).value: r for r in (4, 5)}
    # valid answer canonicalized to the template's exact spelling
    assert ws.cell(rows["AAA"], hdr["Prominent Colour"]).value == "White"
    # invalid answer is NOT written and is reported back for re-prompting
    assert ws.cell(rows["BBB"], hdr["Prominent Colour"]).value in (None, "")
    assert summary["rejected"]["BBB"][0]["field"] == "Prominent Colour"


def test_correct_fills_brand_and_address(tmp_path):
    from src.myntra.error_reader import RowError
    from src.myntra.corrector import correct

    template = read_template(TEMPLATE)
    constants = {
        "brand": "Ijor Ethnic Partners",
        "Manufacturer Name and Address with Pincode": "Ijor, Faridabad, 121006",
        "Packer Name and Address with Pincode": "Ijor, Faridabad, 121006",
    }
    rows = [
        RowError(row=4, sku="AAA", status="", cells={"vendorSkuCode": "AAA", "brand": ""},
                 issues=[{"category": "brand", "action": "auto_fix", "field": None,
                          "explanation": "brand", "raw": "getBrandCodeFromBrandName"}]),
        RowError(row=5, sku="BBB", status="",
                 cells={"vendorSkuCode": "BBB",
                        "Manufacturer Name and Address with Pincode": ""},
                 issues=[{"category": "address", "action": "auto_fix", "field": None,
                          "explanation": "addr", "raw": "information is incomplete"}]),
    ]
    out = tmp_path / "out.xlsx"
    summary = correct(rows, template, TEMPLATE, constants, {}, set(), str(out))
    assert "brand" in summary["changed"]["AAA"]
    assert "Manufacturer Name and Address with Pincode" in summary["changed"]["BBB"]


def test_image_and_stylegroupid_explain_not_auto(tmp_path):
    p = tmp_path / "resub.xlsx"
    _make_resub(p, [
        {"status": "SKU_VALIDATION_FAILED",
         "message": "For the image column: Front Image, extension is not jpg",
         "cells": {"styleGroupId": "11", "vendorSkuCode": "IMG1",
                   "brand": "Ijor Ethnic Partners", "Front Image": f"{IMG}/IMG1/1.webp"}},
        {"status": "SKU_VALIDATION_FAILED",
         "message": "Style SKU Count mismatch",
         "cells": {"styleGroupId": "12", "vendorSkuCode": "SGI1",
                   "brand": "Ijor Ethnic Partners", "Front Image": f"{IMG}/SGI1/1.jpg"}},
    ])
    rules = load_rules()
    errs = read_errors(str(p), rules)
    plan = plan_corrections(errs)
    # neither is falsely promised as an automatic fix
    assert "IMG1" not in plan["auto"]
    assert "SGI1" not in plan["auto"]
    # both surface as explain-only with a helpful, non-empty explanation
    explained = {e["sku"]: e for e in plan["unknown"]}
    assert {"IMG1", "SGI1"} <= set(explained)
    assert explained["IMG1"].get("explanation")
    assert explained["SGI1"].get("explanation")


def test_correct_from_issues_excludes_explain_only_and_logs(tmp_path):
    from src.myntra.explainer import ExplainedIssue
    from src.myntra.corrector import correct_from_issues
    from src.myntra.correction_log import read_log
    from src.web.settings import LocalJsonStore

    template = read_template(TEMPLATE)
    constants = {"brand": "Ijor Ethnic Partners"}

    def _iss(sku, action, category, cells, explanation="x", field=None):
        return ExplainedIssue(sku=sku, style_id=None, scope="sku",
                              source_type="sku_xlsx", raw_reason="getBrandCodeFromBrandName",
                              explanation=explanation, action=action, field=field,
                              category=category, source="yaml", cells=cells)

    issues = [
        _iss("AAA", "auto_fix", "brand", {"vendorSkuCode": "AAA", "brand": ""}),
        _iss("IMG", "explain_only", "image", {"vendorSkuCode": "IMG"},
             explanation="Reshoot the photo"),
    ]
    log = LocalJsonStore(str(tmp_path / "log.json"))
    out = tmp_path / "out.xlsx"
    summary = correct_from_issues(issues, template, TEMPLATE, constants, {},
                                  str(out), log_store=log, fix_id="fix123")

    assert summary["written"] == 1                       # only AAA written
    assert [m["sku"] for m in summary["manual_needed"]] == ["IMG"]
    assert "brand" in summary["changed"]["AAA"]
    recs = read_log(log)
    assert recs[0]["sku"] == "AAA"
    assert recs[0]["fix_id"] == "fix123"
    assert "brand" in recs[0]["changes"]


def test_correct_from_issues_drops_sku(tmp_path):
    from src.myntra.explainer import ExplainedIssue
    from src.myntra.corrector import correct_from_issues
    from src.myntra.correction_log import read_log
    from src.web.settings import LocalJsonStore

    template = read_template(TEMPLATE)
    constants = {"brand": "Ijor Ethnic Partners"}

    def _iss(sku, action, category, cells, explanation="x", field=None):
        return ExplainedIssue(sku=sku, style_id=None, scope="sku",
                              source_type="sku_xlsx", raw_reason="Seller Sku Code is already registered",
                              explanation=explanation, action=action, field=field,
                              category=category, source="yaml", cells=cells)

    issues = [
        _iss("AAA", "auto_fix", "brand", {"vendorSkuCode": "AAA", "brand": ""}),
        _iss("DUP", "drop_sku", "duplicate", {"vendorSkuCode": "DUP"},
             explanation="Already registered, dropping"),
    ]
    log = LocalJsonStore(str(tmp_path / "log.json"))
    out = tmp_path / "out.xlsx"
    summary = correct_from_issues(issues, template, TEMPLATE, constants, {},
                                  str(out), log_store=log, fix_id="fix123")

    assert summary["written"] == 1                        # only AAA written
    assert summary["dropped"] == ["DUP"]
    assert "DUP" not in summary["changed"]
    recs = read_log(log)
    assert [r["sku"] for r in recs] == ["AAA"]             # dropped SKU logs nothing


def test_correct_from_issues_explicit_drops_authoritative(tmp_path):
    """`drops` (when passed as a set) is authoritative over issue action: a
    drop_sku-flagged SKU absent from `drops` is KEPT; one present is dropped."""
    from src.myntra.explainer import ExplainedIssue
    from src.myntra.corrector import correct_from_issues
    from src.web.settings import LocalJsonStore

    template = read_template(TEMPLATE)
    constants = {"brand": "Ijor Ethnic Partners"}

    def _iss(sku, action, category, cells, explanation="x", field=None):
        return ExplainedIssue(sku=sku, style_id=None, scope="sku",
                              source_type="sku_xlsx", raw_reason="Seller Sku Code is already registered",
                              explanation=explanation, action=action, field=field,
                              category=category, source="yaml", cells=cells)

    issues = [
        _iss("KEEP", "drop_sku", "duplicate", {"vendorSkuCode": "KEEP"},
             explanation="Already registered, dropping"),
        _iss("DROP", "drop_sku", "duplicate", {"vendorSkuCode": "DROP"},
             explanation="Already registered, dropping"),
    ]
    log = LocalJsonStore(str(tmp_path / "log.json"))

    # KEEP's checkbox was unchecked (not in drops) -> must be kept, not dropped.
    out1 = tmp_path / "out1.xlsx"
    summary = correct_from_issues(issues, template, TEMPLATE, constants, {},
                                  str(out1), log_store=log, fix_id="fix1",
                                  drops=set())
    assert "KEEP" not in summary["dropped"]
    assert summary["written"] == 2                        # both rows written

    # DROP's checkbox was checked (in drops) -> must be dropped.
    out2 = tmp_path / "out2.xlsx"
    summary2 = correct_from_issues(issues, template, TEMPLATE, constants, {},
                                   str(out2), log_store=log, fix_id="fix2",
                                   drops={"DROP"})
    assert summary2["dropped"] == ["DROP"]
    assert "DROP" not in summary2["changed"]
    assert "KEEP" not in summary2["dropped"]


def test_regenerate_surface_b_resolves_pins_and_reports_missing(monkeypatch, tmp_path):
    import src.myntra.corrector as corrector
    from src.web.settings import Settings

    # Fake registry: AAA is known (has pins), BBB is unknown.
    monkeypatch.setattr(corrector, "sku_registry_store", lambda s: object())
    monkeypatch.setattr(corrector, "read_registry",
                        lambda store: {"AAA": {"style_group_id": 42, "hsn": "52081120"}})

    captured = {}

    def fake_pipeline(**kwargs):
        captured.update(kwargs)
        return {"filled": str(tmp_path / "myntra_filled.xlsx"),
                "products": 1, "records": [{"sku": "AAA"}]}

    monkeypatch.setattr(corrector, "pipeline_main", fake_pipeline)

    summary = corrector.regenerate_surface_b(["AAA", "BBB"], Settings(), str(tmp_path))
    assert captured["only_skus"] == {"AAA", "BBB"}
    assert captured["style_group_id_by_sku"] == {"AAA": 42}
    assert captured["hsn_by_sku"] == {"AAA": "52081120"}
    assert summary["fixed"] == ["AAA"]
    assert summary["could_not_rebuild"] == ["BBB"]


def test_regenerate_surface_b_whole_sheet_applies_registry_pins(monkeypatch, tmp_path):
    import src.myntra.corrector as corrector
    from src.web.settings import Settings

    # Whole-sheet rebuild (skus=None): the registry's known SKUs must still get
    # their recorded styleGroupId + HSN pinned, not fresh sequential ones.
    monkeypatch.setattr(corrector, "sku_registry_store", lambda s: object())
    monkeypatch.setattr(corrector, "read_registry",
                        lambda store: {"AAA": {"style_group_id": 42, "hsn": "52081120"}})

    captured = {}

    def fake_pipeline(**kwargs):
        captured.update(kwargs)
        return {"filled": str(tmp_path / "myntra_filled.xlsx"),
                "products": 1, "records": [{"sku": "AAA"}]}

    monkeypatch.setattr(corrector, "pipeline_main", fake_pipeline)

    summary = corrector.regenerate_surface_b(None, Settings(), str(tmp_path))
    assert captured["only_skus"] is None
    assert captured["style_group_id_by_sku"] == {"AAA": 42}
    assert captured["hsn_by_sku"] == {"AAA": "52081120"}
    assert summary["could_not_rebuild"] == []
