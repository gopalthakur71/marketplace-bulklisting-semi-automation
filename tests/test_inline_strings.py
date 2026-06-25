import re
import warnings
import zipfile

import openpyxl

from src.myntra.template_reader import read_template
from src.core.models import MappedRow, ImageResult
from src.myntra.fill import fill_template

TEMPLATE = "templates/myntra/Myntra-Sku-Template-2026-06-16.xlsx"


def test_sarees_sheet_uses_inline_strings(tmp_path):
    """Myntra's parser ignores shared strings; the Sarees sheet must be inline."""
    warnings.filterwarnings("ignore")
    t = read_template(TEMPLATE)
    row = MappedRow(sku="S1", cells={"vendorSkuCode": "S1", "brand": "Ijor Ethnic Partners",
                                     "vendorArticleName": "Banarasi Blue Saree"})
    out = tmp_path / "filled.xlsx"
    fill_template(TEMPLATE, t, [(row, ImageResult(sku="S1"))], str(out))

    xml = zipfile.ZipFile(str(out)).read("xl/worksheets/sheet2.xml").decode("utf-8")
    assert 't="s"' not in xml                 # no shared-string references
    assert 't="inlineStr"' in xml             # text is inline

    # Values must survive the conversion and reload cleanly.
    ws = openpyxl.load_workbook(str(out))["Sarees"]
    assert ws.cell(3, 6).value == "brand"     # header intact
    assert ws.cell(t.first_data_row, 6).value == "Ijor Ethnic Partners"
    assert ws.cell(t.first_data_row, 5).value == "Banarasi Blue Saree"
