"""Extract the Myntra template's dropdown vocabulary into a downloadable workbook.

Reads a Myntra bulk-listing template (.xlsx), walks every column's data-validation
list on the entry sheet, resolves each list against the (usually hidden) master-data
sheet it references, and writes `outputs/myntra-attribute-vocabulary.xlsx`:

  * a `Summary` sheet — one row per template column: section, constrained vs
    free-text, source range, and accepted-value count;
  * one sheet per constrained attribute — its full accepted-value list, in order,
    exact duplicates removed, nothing else reformatted.

This is the single source of truth for the normalization dictionaries described in
docs/journal/myntra-attribute-mapping-fix-spec.md (Task 0).

Usage:
    python scripts/build_vocabulary_workbook.py [TEMPLATE.xlsx] [OUTPUT.xlsx]

Defaults: newest Myntra-Sku-Template-*.xlsx in the repo root ->
          outputs/myntra-attribute-vocabulary.xlsx
"""
from __future__ import annotations

import glob
import os
import re
import sys

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Attributes whose value list is huge and irrelevant to our normalization:
# `brand` is validated against Myntra's entire registry (~56k names) but we only
# ever emit "Ijor" -> keep it in the Summary (with count) but omit the values sheet.
OMIT_VALUE_SHEETS = {"brand"}

# Excel forbids these in sheet names, plus a 31-char limit and uniqueness.
_BAD_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")
_RANGE_RE = re.compile(
    r"^(?:'?(?P<sheet>[^'!]+)'?!)?\$?(?P<c1>[A-Z]+)\$?(?P<r1>\d+)"
    r"(?::\$?(?P<c2>[A-Z]+)\$?(?P<r2>\d+))?$"
)


def find_template() -> str:
    # Prefer templates/myntra/, fall back to repo root; newest filename wins.
    hits = sorted(glob.glob(os.path.join(ROOT, "templates", "myntra", "Myntra-Sku-Template-*.xlsx")))
    hits += sorted(glob.glob(os.path.join(ROOT, "Myntra-Sku-Template-*.xlsx")))
    if not hits:
        raise SystemExit("No Myntra-Sku-Template-*.xlsx found in templates/myntra/ or repo root.")
    return hits[-1]


def find_entry_sheet(wb):
    """Entry sheet = the one whose column A holds 'styleId' in the first ~11 rows."""
    for ws in wb.worksheets:
        for r in range(1, 12):
            if ws.cell(r, 1).value == "styleId":
                return ws, r
    raise SystemExit("Could not locate the entry sheet (no 'styleId' header found).")


def section_bands(ws, header_row):
    """Row above the header carries section labels spanning column ranges."""
    band_row = header_row - 1
    marks = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(band_row, c).value
        if v not in (None, ""):
            marks.append((c, str(v).split("(")[0].strip()))
    marks.sort()

    def section_for(col_idx):
        label = ""
        for start, lbl in marks:
            if col_idx >= start:
                label = lbl
            else:
                break
        return label

    return section_for


def dv_by_column(ws, data_row):
    """Map column-letter -> (type, formula1) for validations covering the first data row."""
    out = {}
    for dv in ws.data_validations.dataValidation:
        for rng in dv.sqref.ranges:
            if rng.min_row <= data_row <= rng.max_row:
                for c in range(rng.min_col, rng.max_col + 1):
                    out.setdefault(get_column_letter(c), (dv.type, str(dv.formula1)))
    return out


def resolve_list(wb, formula: str):
    """Resolve a list data-validation formula to its accepted values (in order)."""
    f = (formula or "").strip().strip('"')
    if not f:
        return []
    # Inline comma list e.g.  Yes,No
    if "!" not in f and "$" not in f and "," in f:
        return [v.strip() for v in f.split(",") if v.strip()]
    m = _RANGE_RE.match(f)
    if not m:
        return []
    sheet = m.group("sheet")
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else None
    if ws is None:
        return []
    c1 = column_index_from_string(m.group("c1"))
    r1 = int(m.group("r1"))
    r2 = int(m.group("r2")) if m.group("r2") else r1
    if r2 < r1:  # reversed/empty range e.g. $AJ$2:$AJ$1
        return []
    values = []
    for r in range(r1, r2 + 1):
        v = ws.cell(r, c1).value
        if v is None:
            continue
        s = str(v).strip()
        if s:
            values.append(s)
    # remove exact duplicates, preserve first-seen order
    return list(dict.fromkeys(values))


def safe_sheet_name(name, used):
    base = _BAD_SHEET_CHARS.sub(" ", name).strip()[:31] or "Sheet"
    candidate = base
    i = 1
    while candidate.lower() in used:
        suffix = f" ({i})"
        candidate = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(candidate.lower())
    return candidate


def build(template_path, out_path):
    wb = openpyxl.load_workbook(template_path, data_only=True, keep_links=False)
    entry, header_row = find_entry_sheet(wb)
    data_row = header_row + 1
    section_for = section_bands(entry, header_row)
    dvmap = dv_by_column(entry, data_row)

    columns = []  # (col_letter, header, section, source, values|None)
    for c in range(1, entry.max_column + 1):
        header = entry.cell(header_row, c).value
        if header in (None, ""):
            continue
        letter = get_column_letter(c)
        dv = dvmap.get(letter)
        if dv and dv[0] == "list":
            values = resolve_list(wb, dv[1])
            source = dv[1].strip().strip('"')
        else:
            values = None
            source = ""
        columns.append((letter, str(header), section_for(c), source, values))

    out_wb = openpyxl.Workbook()
    summary = out_wb.active
    summary.title = "Summary"
    summary.append(["Column", "Attribute (header)", "Section", "Type",
                    "Accepted values", "Source range", "Sheet"])
    for cell in summary[1]:
        cell.font = openpyxl.styles.Font(bold=True)
    summary.freeze_panes = "A2"

    used = {"summary"}
    per_attr = []  # (sheet_name, header, values)
    for letter, header, section, source, values in columns:
        if values is None:
            summary.append([letter, header, section, "free-text", "", "", ""])
        elif header in OMIT_VALUE_SHEETS:
            summary.append([letter, header, section, "dropdown",
                            len(values), source, "(list omitted)"])
        else:
            sheet_name = safe_sheet_name(header, used)
            per_attr.append((sheet_name, header, values))
            summary.append([letter, header, section, "dropdown",
                            len(values), source, sheet_name])

    # widths
    for col, w in zip("ABCDEFG", (8, 34, 26, 11, 16, 26, 24)):
        summary.column_dimensions[col].width = w

    for sheet_name, header, values in per_attr:
        ws = out_wb.create_sheet(sheet_name)
        ws.append([header])
        ws["A1"].font = openpyxl.styles.Font(bold=True)
        ws.freeze_panes = "A2"
        ws.column_dimensions["A"].width = min(60, max(20, len(header) + 4))
        for v in values:
            ws.append([v])

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    out_wb.save(out_path)

    constrained = [c for c in columns if c[4] is not None]
    free = [c for c in columns if c[4] is None]
    print(f"Template : {os.path.relpath(template_path, ROOT)}")
    print(f"Entry    : {entry.title!r} (header row {header_row})")
    print(f"Columns  : {len(columns)} total | {len(constrained)} dropdown | {len(free)} free-text")
    print(f"Output   : {os.path.relpath(out_path, ROOT)}\n")
    print("Dropdown attributes (value counts):")
    for _, header, _, _, values in constrained:
        note = "  <-- EMPTY" if not values else ""
        print(f"  {header[:36]:37} {len(values):>6}{note}")


if __name__ == "__main__":
    template = sys.argv[1] if len(sys.argv) > 1 else find_template()
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
        ROOT, "outputs", "myntra-attribute-vocabulary.xlsx")
    build(template, out)
