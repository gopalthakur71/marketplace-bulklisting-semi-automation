import os
import re
import shutil
import tempfile
import warnings
import zipfile

import openpyxl

IMAGE_COLUMNS = ["Front Image", "Side Image", "Back Image", "Detail Angle",
                 "Look Shot Image", "Additional Image 1", "Additional Image 2"]

SHEET_SAREES_NAME = "Sarees"


def _sheet_xml_name(xlsx_path, sheet_title):
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    idx = wb.sheetnames.index(sheet_title)
    wb.close()
    return f"xl/worksheets/sheet{idx + 1}.xml"


def _parse_shared_strings(xlsx_path):
    """Return shared strings as a list (index -> raw, already-XML-escaped text)."""
    try:
        with zipfile.ZipFile(xlsx_path) as z:
            xml = z.read("xl/sharedStrings.xml").decode("utf-8")
    except KeyError:
        return []
    out = []
    for si in re.findall(r"<si>(.*?)</si>", xml, re.S):
        # Concatenate all <t>..</t> runs (handles plain and rich-text strings).
        parts = re.findall(r"<t[^>]*>(.*?)</t>", si, re.S)
        out.append("".join(parts))
    return out


def _shared_to_inline(out_path, sheet_xml_name):
    """Convert shared-string cells (t="s") in one sheet to inline strings
    (t="inlineStr"). Myntra's upload parser does not resolve shared strings, so
    text — including the column headers — must be embedded inline."""
    strings = _parse_shared_strings(out_path)
    if not strings:
        return

    def repl(m):
        before, after, idx = m.group(1), m.group(2), int(m.group(3))
        text = strings[idx] if idx < len(strings) else ""
        return (f'<c {before}t="inlineStr"{after}>'
                f'<is><t xml:space="preserve">{text}</t></is></c>')

    # <c r=".." s=".." t="s"><v>N</v></c>  (t may sit anywhere in the attributes)
    pattern = re.compile(r'<c ([^>]*?)t="s"([^>]*?)>\s*<v>(\d+)</v>\s*</c>')

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    with zipfile.ZipFile(out_path) as zin, \
            zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_xml_name:
                data = pattern.sub(repl, data.decode("utf-8")).encode("utf-8")
            zout.writestr(item, data)
    shutil.move(tmp_path, out_path)


def _extract_validation_ext(template_path, sheet_xml_name):
    """Return the self-contained <ext>..</ext> x14 dataValidations block from the
    template's Sarees sheet, with xr:uid attributes stripped (so it needs no xr ns)."""
    with zipfile.ZipFile(template_path) as z:
        xml = z.read(sheet_xml_name).decode("utf-8")
    m = re.search(r"<ext\b[^>]*>\s*<x14:dataValidations.*?</x14:dataValidations>\s*</ext>", xml, re.S)
    if not m:
        return None
    block = m.group(0)
    block = re.sub(r'\s+xr:uid="[^"]*"', "", block)
    return block


def _inject_validations(out_path, sheet_xml_name, ext_block):
    """Insert the x14 validation ext block back into the saved workbook's sheet XML."""
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)
    with zipfile.ZipFile(out_path) as zin, zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_xml_name:
                xml = data.decode("utf-8")
                if "</extLst>" in xml:
                    xml = xml.replace("</extLst>", ext_block + "</extLst>", 1)
                else:
                    xml = xml.replace("</worksheet>", f"<extLst>{ext_block}</extLst></worksheet>", 1)
                data = xml.encode("utf-8")
            zout.writestr(item, data)
    shutil.move(tmp_path, out_path)


def fill_template(template_path, template, rows, out_path, preserve_dropdowns=False):
    """Fill the Sarees sheet and save.

    preserve_dropdowns re-injects the template's x14 dropdown validations into the
    saved file. This is OFF by default: Myntra's upload parser (Apache POI) rejects
    the re-injected extension XML ("Error while reading and validating the input
    file"), and the dropdowns are not needed for upload — only for manual editing.
    """
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(template_path)
    ws = wb[SHEET_SAREES_NAME]

    r = template.first_data_row
    for mapped, images in rows:
        for header, value in mapped.cells.items():
            col = template.col_index_by_header.get(header)
            if col:
                ws.cell(row=r, column=col, value=value)
        # Myntra ingests images by URL, so write the validated CDN URLs (not local
        # filenames) into the image columns. Falls back to local basenames only if
        # no URLs were tracked.
        image_values = images.passed_urls or [os.path.basename(p) for p in images.passed]
        for header, value in zip(IMAGE_COLUMNS, image_values):
            col = template.col_index_by_header.get(header)
            if col:
                ws.cell(row=r, column=col, value=value)
        r += 1

    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)

    # Myntra's upload parser does not resolve shared strings; convert the Sarees
    # sheet's text cells (including headers) to inline strings.
    sarees_xml = _sheet_xml_name(template_path, SHEET_SAREES_NAME)
    _shared_to_inline(out_path, sarees_xml)

    # Re-inject the dropdown validations openpyxl dropped on save (manual-edit copy
    # only — breaks Myntra's upload parser, so off by default).
    if preserve_dropdowns:
        sheet_xml = _sheet_xml_name(template_path, SHEET_SAREES_NAME)
        ext_block = _extract_validation_ext(template_path, sheet_xml)
        if ext_block:
            _inject_validations(out_path, sheet_xml, ext_block)
