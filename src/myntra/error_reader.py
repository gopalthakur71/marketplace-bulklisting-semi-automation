from dataclasses import dataclass, field

import openpyxl
import yaml

ERROR_COLUMNS = {"STATUS", "SYSTEM ERROR MESSAGE"}
HEADER_ROW = 3
FIRST_DATA_ROW = 4


@dataclass
class RowError:
    row: int
    sku: str
    status: str
    cells: dict          # standard header -> value (error columns stripped)
    issues: list = field(default_factory=list)


def load_rules(path="config/myntra/error_rules.yaml"):
    with open(path, encoding="utf-8") as fh:
        return yaml.safe_load(fh)


def classify(message, rules):
    msg = (message or "").strip()
    low = msg.lower()
    for rule in rules.get("rules", []):
        if str(rule["match"]).lower() in low:
            return {"category": rule["category"], "action": rule["action"],
                    "explanation": rule["explanation"], "field": rule.get("field"),
                    "raw": msg}
    unk = rules["unknown"]
    return {"category": unk["category"], "action": unk["action"],
            "explanation": unk["explanation"], "field": None, "raw": msg}


def read_errors(path, rules, sheet="Sarees"):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    headers = [ws.cell(HEADER_ROW, c).value for c in range(1, ws.max_column + 1)]
    out = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        status = ws.cell(r, 1).value
        message = ws.cell(r, 2).value
        if status is None and message is None:
            continue
        cells = {}
        for c, h in enumerate(headers, start=1):
            if h in ERROR_COLUMNS or h is None:
                continue
            v = ws.cell(r, c).value
            cells[h] = None if v is None else str(v)
        issues = [classify(m, rules) for m in str(message or "").split(";") if m.strip()]
        out.append(RowError(row=r, sku=cells.get("vendorSkuCode") or "",
                             status=str(status or ""), cells=cells, issues=issues))
    return out
