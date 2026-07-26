import re
import zipfile
import warnings

import openpyxl
from openpyxl.utils import column_index_from_string

from src.core.models import TemplateInfo

SHEET_SAREES_NAME = "Sarees"
MASTERDATA_NAME = "masterdata"


def _find_sheet_xml_name(xlsx_path, sheet_title):
    """Return the worksheets/sheetN.xml path for a given sheet title."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    idx = wb.sheetnames.index(sheet_title)  # 0-based
    wb.close()
    return f"xl/worksheets/sheet{idx + 1}.xml"


def _parse_x14_validations(xlsx_path, sheet_xml_name):
    """Return list of (col_index, masterdata_col_index, first_row, last_row)."""
    with zipfile.ZipFile(xlsx_path) as z:
        xml = z.read(sheet_xml_name).decode("utf-8")
    out = []
    for block in re.findall(r"<x14:dataValidation\b.*?</x14:dataValidation>", xml, re.S):
        fm = re.search(r"<xm:f>(.*?)</xm:f>", block, re.S)
        sq = re.search(r"<xm:sqref>(.*?)</xm:sqref>", block, re.S)
        if not (fm and sq):
            continue
        col_letter = re.match(r"([A-Z]+)", sq.group(1)).group(1)
        col_index = column_index_from_string(col_letter)
        m = re.search(r"masterdata!\$([A-Z]+)\$(\d+):\$([A-Z]+)\$(\d+)", fm.group(1))
        if not m:
            continue
        md_col = column_index_from_string(m.group(1))
        out.append((col_index, md_col, int(m.group(2)), int(m.group(4))))
    return out


_MD_RANGE = re.compile(
    r"(?:'?(?P<sheet>[^'!]+)'?!)?\$?(?P<c1>[A-Z]+)\$?(?P<r1>\d+)"
    r"(?::\$?(?P<c2>[A-Z]+)\$?(?P<r2>\d+))?$"
)


def _plain_list_validations(ws, data_row):
    """(col_index, formula1) for each plain list validation covering data_row.
    First validation seen per column wins."""
    out = {}
    for dv in ws.data_validations.dataValidation:
        if dv.type != "list":
            continue
        for rng in dv.sqref.ranges:
            if rng.min_row <= data_row <= rng.max_row:
                for c in range(rng.min_col, rng.max_col + 1):
                    out.setdefault(c, str(dv.formula1))
    return out


def _resolve_validation_values(wb, formula):
    """Resolve a list-validation formula to accepted values, in order, exact dups
    removed. Handles inline 'A,B,C' lists and masterdata!$X$r0:$X$r1 ranges."""
    f = (formula or "").strip().strip('"')
    if not f:
        return []
    if "!" not in f and "$" not in f and "," in f:            # inline list
        return list(dict.fromkeys(v.strip() for v in f.split(",") if v.strip()))
    m = _MD_RANGE.match(f)
    if not m:
        return []
    sheet = m.group("sheet")
    if not sheet or sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    col = column_index_from_string(m.group("c1"))
    r1 = int(m.group("r1"))
    r2 = int(m.group("r2")) if m.group("r2") else r1
    if r2 < r1:
        return []
    values = []
    for r in range(r1, r2 + 1):
        v = ws.cell(row=r, column=col).value
        if v not in (None, ""):
            values.append(str(v).strip())
    return list(dict.fromkeys(values))


def read_template(path):
    warnings.filterwarnings("ignore")
    wb = openpyxl.load_workbook(path)
    ws = wb[SHEET_SAREES_NAME]
    md = wb[MASTERDATA_NAME]

    # detect header row: row whose column 1 == 'styleId'
    header_row = next(r for r in range(1, 11) if ws.cell(row=r, column=1).value == "styleId")
    max_col = ws.max_column
    headers = [ws.cell(row=header_row, column=c).value for c in range(1, max_col + 1)]
    col_index_by_header = {h: i + 1 for i, h in enumerate(headers) if h not in (None, "")}

    data_row = header_row + 1
    plain = _plain_list_validations(ws, data_row)
    vocab_by_header = {}
    if plain:
        for col_index, formula in plain.items():
            header = headers[col_index - 1] if col_index - 1 < len(headers) else None
            if not header:
                continue
            values = _resolve_validation_values(wb, formula)
            if values:
                vocab_by_header[header] = values
    else:
        sheet_xml = _find_sheet_xml_name(path, SHEET_SAREES_NAME)
        for col_index, md_col, r0, r1 in _parse_x14_validations(path, sheet_xml):
            header = headers[col_index - 1] if col_index - 1 < len(headers) else None
            if not header:
                continue
            values = []
            for r in range(r0, r1 + 1):
                v = md.cell(row=r, column=md_col).value
                if v not in (None, ""):
                    values.append(str(v).strip())
            vocab_by_header[header] = values

    wb.close()
    return TemplateInfo(
        headers=[h for h in headers if h not in (None, "")],
        header_row=header_row,
        first_data_row=header_row + 1,
        col_index_by_header=col_index_by_header,
        vocab_by_header=vocab_by_header,
    )
