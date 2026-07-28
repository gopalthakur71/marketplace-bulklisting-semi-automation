"""Reconstruct Myntra's auto-generated title/design-details from a filled row's
attributes (approximate, for the in-app preview) and read a filled workbook.

Myntra generates the customer-facing title and 'Design Details' prose from the
attribute columns; the exact wording is not something we can guarantee, so these
reconstructions are labelled approximate in the UI. Specifications shown alongside
are exact (they are the values the user entered)."""
import warnings

import openpyxl

SHEET_SAREES_NAME = "Sarees"

# Title order, reverse-engineered from live IJOR listings (2026-07-24). Colour is
# NOT in the saree title. Only attributes that are set are included.
_TITLE_ORDER = ["Print or Pattern Type", "Ornamentation", "Saree Fabric", "Type"]


# Myntra displays metallic colours with a "-Toned" suffix ("Gold" -> "Gold-Toned").
# Confirmed on SKU 164SDE226RPPG, whose Green + Gold published as "Green and
# Gold-Toned Khadi sarees". Members are the metallics in the Prominent Colour
# vocabulary; extend as more live listings confirm others.
_TONED_COLOURS = frozenset({
    "Gold", "Silver", "Bronze", "Copper", "Rose Gold", "Metallic", "Champagne",
})


def is_set(value):
    """True unless the value is None, blank, or the literal 'NA' (case-insensitive)."""
    if value is None:
        return False
    s = str(value).strip()
    return s != "" and s.upper() != "NA"


def reconstruct_title(attrs):
    parts = [str(attrs.get(h)).strip() for h in _TITLE_ORDER if is_set(attrs.get(h))]
    parts.append("Saree")
    title = " ".join(parts)
    if is_set(attrs.get("Blouse Fabric")):
        title += " With Unstitched Blouse Piece"
    return title


def _colour_display(value):
    """Myntra's display name for one colour, e.g. "Gold" -> "Gold-Toned"."""
    c = str(value).strip()
    return c + "-Toned" if c in _TONED_COLOURS else c


def _colour_phrase(attrs):
    """The colour half of Design Details line 1. Myntra joins the prominent and
    second prominent colours with "and"; a third colour has never been observed
    in the prose, so it is left out."""
    first, second = attrs.get("Prominent Colour"), attrs.get("Second Prominent Colour")
    if not is_set(first):
        return ""
    phrase = _colour_display(first)
    if is_set(second):
        phrase += " and " + _colour_display(second)
    return phrase


def reconstruct_design_details(attrs):
    lines = []
    typ = attrs.get("Type")
    l1 = _colour_phrase(attrs)
    if l1:
        if is_set(typ):
            l1 += " " + str(typ).strip()
        lines.append(l1 + " sarees")
    pattern, border = attrs.get("Pattern"), attrs.get("Border")
    if is_set(pattern) or is_set(border):
        p = str(pattern).strip() if is_set(pattern) else ""
        # Myntra renders this as "<Border> Border border" — it appends both the
        # capitalised and the lowercase word to a Border value that is already a
        # clean single term ("Woven Design" -> "Woven Design Border border").
        # That doubling is Myntra's, not ours; we mirror it so the preview matches
        # the live page instead of quietly showing a tidier string.
        b = f"with {str(border).strip()} Border border" if is_set(border) else ""
        lines.append(f"{p} saree {b}".replace("  ", " ").strip())
    orn = attrs.get("Ornamentation")
    if is_set(orn):
        lines.append(f"Has {str(orn).strip()} detail")
    return lines


def missing_attributes(attrs, user_filled):
    return [h for h in user_filled if not is_set(attrs.get(h))]


def build_card(attrs, user_filled):
    """The one place a preview card is assembled, so the upload preview and the
    in-app live preview can never drift apart."""
    return {
        "sku": attrs.get("vendorSkuCode") or attrs.get("SKUCode") or "",
        "title": reconstruct_title(attrs),
        "design_details": reconstruct_design_details(attrs),
        "specs": [(h, attrs.get(h)) for h in user_filled],
        "missing": missing_attributes(attrs, user_filled),
    }


def read_filled_rows(xlsx_path, template):
    """One {header: value_or_None} dict per data row that carries a vendorSkuCode."""
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        try:
            ws = wb[SHEET_SAREES_NAME]
            sku_col = template.col_index_by_header.get("vendorSkuCode")
            rows = []
            for r in range(template.first_data_row, ws.max_row + 1):
                if sku_col is None or not is_set(ws.cell(row=r, column=sku_col).value):
                    continue
                cells = {}
                for header, col in template.col_index_by_header.items():
                    v = ws.cell(row=r, column=col).value
                    cells[header] = None if v is None else str(v).strip()
                rows.append(cells)
        finally:
            wb.close()
    return rows
