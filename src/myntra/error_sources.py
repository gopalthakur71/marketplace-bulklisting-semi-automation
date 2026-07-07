import csv
import os
from dataclasses import dataclass

import openpyxl

from src.myntra.error_reader import read_errors, load_rules

_XLSX_HEADERS = {"STATUS", "SYSTEM ERROR MESSAGE"}
_SHEET_CSV_HEADERS = {"row no", "status", "system error message"}
_LISTINGS_HEADERS = {"style status", "seller sku code", "onhold reason"}


@dataclass
class ErrorItem:
    sku: str | None
    style_id: str | None
    source_type: str          # sku_xlsx | sheet_csv | listings_report
    scope: str                # sku | sheet
    raw_reason: str
    cells: dict | None


def _xlsx_error_sheet(path):
    """First worksheet whose header row (scanning rows 1..6) holds both error
    columns -> (sheet_name, header_row). Fixes the old hardcoded sheet='Sarees'.

    Uses read_only=False (matching error_reader.read_errors) because several
    real Myntra rejection files carry a stale <dimension> tag (e.g. reporting
    max_row=1, max_col=1 for a sheet that actually has 5 rows x 82 cols) --
    read_only mode trusts that stale metadata and iter_rows() silently returns
    nothing, so detection would wrongly report None for real fixtures.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        for ws in wb.worksheets:
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True),
                                    start=1):
                vals = {str(v).strip() for v in row if v is not None}
                if _XLSX_HEADERS <= vals:
                    return ws.title, i
        return None, None
    finally:
        wb.close()


def _csv_header(path):
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for row in csv.reader(fh):
            return [(c or "").strip() for c in row]
    return []


def detect_format(path):
    """(source_type | None, user_facing_reason). Extension gate, then content
    fingerprint by column presence (spec §4)."""
    ext = os.path.splitext(path)[1].lower()
    if ext not in (".xlsx", ".csv"):
        return None, "Please upload a Myntra rejection .xlsx or .csv file."
    try:
        if ext == ".xlsx":
            sheet, _ = _xlsx_error_sheet(path)
            if sheet:
                return "sku_xlsx", ""
            return None, ("This doesn't look like a Myntra rejection — please upload "
                          "the rejection file or the downloaded Listings Report.")
        header = {h.lower() for h in _csv_header(path)}
        if _SHEET_CSV_HEADERS <= header:
            return "sheet_csv", ""
        if _LISTINGS_HEADERS <= header:
            return "listings_report", ""
        return None, ("This doesn't look like a Myntra rejection or Listings Report — "
                      "please upload the rejection file or the downloaded Listings Report.")
    except Exception:
        return None, "Couldn't read this file."


def _read_sku_xlsx(path, rules):
    sheet, _ = _xlsx_error_sheet(path)
    items = []
    for re_ in read_errors(path, rules, sheet=sheet):
        for issue in re_.issues:
            items.append(ErrorItem(
                sku=re_.sku or None,
                style_id=re_.cells.get("styleId") or re_.cells.get("styleGroupId"),
                source_type="sku_xlsx", scope="sku",
                raw_reason=issue["raw"], cells=re_.cells))
    return items


def _rows_lower(path):
    with open(path, newline="", encoding="utf-8-sig") as fh:
        for row in csv.DictReader(fh):
            yield {(k or "").strip().lower(): (v if v is not None else "") for k, v in row.items()}


def _read_sheet_csv(path):
    # Whole-sheet rejection: one message per row, NOT split on ';' (the trailing
    # ';failed while validation:null' is noise, not a separate error).
    items = []
    for rec in _rows_lower(path):
        msg = str(rec.get("system error message") or "").strip()
        if msg:
            items.append(ErrorItem(sku=None, style_id=None, source_type="sheet_csv",
                                   scope="sheet", raw_reason=msg, cells=None))
    return items


def _read_listings_report(path):
    items = []
    for rec in _rows_lower(path):
        reason = str(rec.get("onhold reason") or "").strip()
        if not reason:
            continue  # live/OK rows carry no onhold reason
        items.append(ErrorItem(
            sku=(rec.get("seller sku code") or None),
            style_id=(rec.get("style id") or None),
            source_type="listings_report", scope="sku",
            raw_reason=reason, cells=None))
    return items


def read_error_file(path, rules=None):
    """Detect the format and return a normalized ErrorItem list. Unknown formats
    return [] — the caller uses detect_format() for the user-facing reason."""
    rules = rules or load_rules()
    src, _ = detect_format(path)
    if src == "sku_xlsx":
        return _read_sku_xlsx(path, rules)
    if src == "sheet_csv":
        return _read_sheet_csv(path)
    if src == "listings_report":
        return _read_listings_report(path)
    return []
