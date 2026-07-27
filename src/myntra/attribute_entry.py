"""The attributes the seller decides by hand (colour, fabric, type, border, ...).

The pipeline never guesses these; they are offered as dropdowns whose options come
strictly from the Myntra template's own vocabulary, and are written into the built
workbook only after an exact-membership check."""
import os
import warnings

import openpyxl
import yaml

from src.myntra.fill import SHEET_SAREES_NAME, sheet_xml_name, shared_to_inline
from src.myntra.preview import is_set

CONFIG_DIR = os.path.join("config", "myntra")

# Free-text and mandatory in the Myntra template, but derived — never typed. See
# docs/superpowers/specs/2026-07-27-brand-colour-auto-fill-design.md.
BRAND_COLOUR_HEADER = "Brand Colour (Remarks)"

# Used only if rules.yaml somehow lacks the key; the YAML is the source of truth.
FALLBACK_USER_FILLED = [
    "Prominent Colour", "Second Prominent Colour", "Third Prominent Colour",
    "Saree Fabric", "Blouse Fabric", "Type", "Ornamentation", "Border",
    "Pattern", "Print or Pattern Type", "Wash Care", "Usage"]


class AttributeValueError(Exception):
    """A submitted value is not an exact member of that column's Myntra vocabulary."""


def user_filled_attributes(config_dir=CONFIG_DIR):
    with open(os.path.join(config_dir, "rules.yaml"), encoding="utf-8") as fh:
        rules = yaml.safe_load(fh) or {}
    return rules.get("user_filled_attributes") or list(FALLBACK_USER_FILLED)


def attribute_vocab(template, columns):
    """{column: [accepted values]} straight from the template. Nothing is added."""
    return {c: list(template.vocab_by_header.get(c) or []) for c in columns}


def validate_submitted(values, vocab):
    """Blank -> None (clears the cell). Non-blank must be exactly in vocab, else raise."""
    out = {}
    for column, value in values.items():
        if column not in vocab:
            raise AttributeValueError(f"Unknown attribute column: {column}")
        if value is None or str(value).strip() == "":
            out[column] = None
            continue
        v = str(value).strip()
        if v not in vocab[column]:
            raise AttributeValueError(
                f"{column}: '{v}' is not one of Myntra's accepted values")
        out[column] = v
    return out


def derive_brand_colour(values):
    """Brand Colour (Remarks) follows Prominent Colour, lowercased.

    Myntra requires this free-text column but derives nothing itself, so a blank
    one gets the row rejected ('Brand Colour (Remarks) cannot be null') and the
    fix flow ends up mirroring the colour in anyway. Deriving it at save time
    closes that round-trip. `NA` is a real vocabulary value meaning 'no colour
    stated', so it mirrors to nothing rather than the string 'na'."""
    colour = values.get("Prominent Colour")
    return str(colour).strip().lower() if is_set(colour) else None


class SkuMismatchError(Exception):
    """A submitted row ordinal does not carry the SKU the form claimed for it."""


def write_attributes(xlsx_path, template, entries):
    """Write the user-chosen attributes into an already-built workbook, in place.

    entries: [{"ordinal": int, "sku": str, "values": {header: value_or_None}}]
    A None value blanks the cell, so re-saving is idempotent and a cleared dropdown
    really clears. Every entry is verified before anything is written."""
    warnings.filterwarnings("ignore")
    sku_col = template.col_index_by_header.get("vendorSkuCode")
    if sku_col is None:
        raise SkuMismatchError("Template has no vendorSkuCode column")

    wb = openpyxl.load_workbook(xlsx_path)
    try:
        ws = wb[SHEET_SAREES_NAME]
        # Pass 1: verify every target row before touching any cell.
        for e in entries:
            r = template.first_data_row + int(e["ordinal"])
            actual = ws.cell(row=r, column=sku_col).value
            actual = "" if actual is None else str(actual).strip()
            if actual != str(e["sku"]).strip():
                raise SkuMismatchError(
                    f"Row {r} holds SKU '{actual}', not '{e['sku']}' — "
                    "the sheet changed since the screen was opened")
        # Pass 2: write.
        for e in entries:
            r = template.first_data_row + int(e["ordinal"])
            for header, value in e["values"].items():
                col = template.col_index_by_header.get(header)
                if col is None:
                    continue
                ws.cell(row=r, column=col).value = value
        wb.save(xlsx_path)
    finally:
        wb.close()

    # openpyxl re-writes text as shared strings; Myntra's parser cannot resolve
    # them, so re-apply fill.py's inline conversion (see fill.fill_template).
    shared_to_inline(xlsx_path, sheet_xml_name(xlsx_path, SHEET_SAREES_NAME))
    return len(entries)
